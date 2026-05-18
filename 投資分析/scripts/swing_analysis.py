"""
短線分析 v9 draft — 自動化參數版（省去 WebSearch）
新增：自動抓 TAIEX 算 stock_rs、自動抓融資餘額算 margin_change
觸發：run("6257","矽格", catalyst="AI封測需求", catalyst_quality="強")
"""
import time, warnings, sys, os, re
import numpy as np
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from matplotlib.backends.backend_pdf import PdfPages

warnings.filterwarnings('ignore')
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from matplotlib.patches import FancyBboxPatch, Rectangle

plt.rcParams['font.family'] = ['PingFang HK','STHeiti','Arial Unicode MS','DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

BG       = '#0d1117'
PANEL    = '#161b22'
PANEL2   = '#1c2128'
WINE     = '#922B21'
WINE_LT  = '#C0392B'
WINE_DK  = '#641E16'
GOLD     = '#D4A843'
GOLD_DIM = '#A68532'
GREEN    = '#2EA043'
RED      = '#DA3633'
BLUE     = '#4493F8'
WHITE    = '#E6E6E6'
GRAY     = '#8B949E'
GRAY_DK  = '#484F58'
PURPLE   = '#A371F7'
TEAL     = '#3FB950'
ORANGE   = '#D29922'

SESSION = requests.Session()
SESSION.headers.update({'User-Agent':'Mozilla/5.0'})

def get_json(url, t=12):
    try: return SESSION.get(url,timeout=t).json()
    except: return {}
def get_text(url, t=25):
    try: return SESSION.get(url,timeout=t).text
    except: return ""
def last_n_wd(n=5):
    d=[]; dt=datetime.today()
    while len(d)<n:
        if dt.weekday()<5: d.append(dt.strftime('%Y%m%d'))
        dt-=timedelta(days=1)
    return d

def sma_np(a,n):
    o=np.full(len(a),np.nan); cs=np.cumsum(a); o[n-1:]=(cs[n-1:]-np.concatenate([[0],cs[:-n]]))/n; return o
def ema_np(a,n):
    k=2/(n+1); o=np.empty(len(a)); o[0]=a[0]
    for i in range(1,len(a)): o[i]=a[i]*k+o[i-1]*(1-k)
    return o
def rsi_np(c,n=14):
    d=np.diff(c.astype(float)); o=np.full(len(c),np.nan)
    for i in range(n,len(c)):
        g=np.where(d[i-n:i]>0,d[i-n:i],0.); l=np.where(d[i-n:i]<0,-d[i-n:i],0.)
        al=l.mean(); o[i]=100 if al==0 else 100-100/(1+g.mean()/al)
    return o
def kd_np(H,L,C,n=9):
    K=np.empty(len(C)); D=np.empty(len(C)); K[0]=D[0]=50.
    for i in range(1,len(C)):
        lo=L[max(0,i-n+1):i+1].min(); hi=H[max(0,i-n+1):i+1].max()
        rsv=50. if hi==lo else (C[i]-lo)/(hi-lo)*100
        K[i]=K[i-1]*2/3+rsv/3; D[i]=D[i-1]*2/3+K[i]/3
    return K,D
def macd_np(c):
    e12=ema_np(c,12); e26=ema_np(c,26); m=e12-e26; s=ema_np(m,9); return m,s,m-s
def boll_np(c,n=20,k=2):
    mid=sma_np(c,n); rs=np.full(len(c),np.nan)
    for i in range(n-1,len(c)): rs[i]=c[i-n+1:i+1].std()
    return mid+k*rs, mid, mid-k*rs
def atr_np(H,L,C,n=10):
    tr=np.concatenate([[H[0]-L[0]],np.maximum(H[1:]-L[1:],np.maximum(np.abs(H[1:]-C[:-1]),np.abs(L[1:]-C[:-1])))])
    return tr[-n:].mean()

def fetch_twse(code,ym):
    d=get_json(f"https://www.twse.com.tw/exchangeReport/STOCK_DAY?response=json&date={ym}01&stockNo={code}")
    return d.get('data',[]) if d.get('stat')=='OK' else []
def fetch_tpex(code,ym):
    """上櫃股 OHLCV（TPEX 新版 API）— 欄位：日期/張數/仟元/開/高/低/收/漲跌/筆"""
    y_ad=f"{ym[:4]}/{ym[4:]}/01"
    d=get_json(f"https://www.tpex.org.tw/www/zh-tw/afterTrading/tradingStock?date={y_ad}&code={code}")
    rows=[]
    for t in d.get('tables',[]):
        for r in t.get('data',[]):
            rows.append(r)
    return rows
def parse_tpex_ohlcv(rows):
    """解析 TPEX：[日期, 張數, 仟元, 開, 高, 低, 收, 漲跌, 筆]"""
    dt=[]; O=[]; H=[]; L=[]; C=[]; V=[]
    for r in rows:
        try:
            dt.append(str(r[0]).strip())
            V.append(float(str(r[1]).replace(',','')))  # 已經是張數
            O.append(float(str(r[3]).replace(',',''))); H.append(float(str(r[4]).replace(',','')))
            L.append(float(str(r[5]).replace(',',''))); C.append(float(str(r[6]).replace(',','')))
        except: pass
    return dt,np.array(O),np.array(H),np.array(L),np.array(C),np.array(V)
def fetch_taiex(ym):
    """取加權指數月資料 → 返回 [(date_str, close), ...]"""
    d=get_json(f"https://www.twse.com.tw/exchangeReport/FMTQIK?response=json&date={ym}01")
    if d.get('stat')!='OK': return []
    out=[]
    for r in d.get('data',[]):
        try: out.append((r[0].strip(), float(r[4].replace(',',''))))
        except: pass
    return out
def fetch_margin(code, date):
    """取個股融資餘額（張）"""
    d=get_json(f"https://www.twse.com.tw/exchangeReport/MI_MARGN?response=json&date={date}&selectType=ALL")
    if d.get('stat')!='OK': return None
    for r in d.get('data',[]):
        if str(r[0]).strip()==code:
            try: return int(str(r[6]).strip().replace(',',''))  # 融資今日餘額（張）
            except: pass
    return None
def parse_ohlcv(rows):
    dt=[]; O=[]; H=[]; L=[]; C=[]; V=[]
    for r in rows:
        try:
            dt.append(r[0].strip()); V.append(float(r[1].replace(',',''))/1000)
            O.append(float(r[3].replace(',',''))); H.append(float(r[4].replace(',','')))
            L.append(float(r[5].replace(',',''))); C.append(float(r[6].replace(',','')))
        except: pass
    return dt,np.array(O),np.array(H),np.array(L),np.array(C),np.array(V)
def fetch_five91(code):
    raw=get_text(f"https://five91.onrender.com/stock/{code}",28)
    # 先從 title 抓公司名（strip 前）
    name=None
    mt=re.search(r'<title[^>]*>([^<]+)</title>',raw,re.IGNORECASE)
    if mt:
        mn=re.search(r'\d{4}\s+([\u4e00-\u9fff\w]+)',mt.group(1))
        if mn: name=mn.group(1)
    txt=re.sub(r'<[^>]+>',' ',raw)  # 清除 HTML tags 再解析
    if not name:
        mn=re.search(rf'{code}\s+([\u4e00-\u9fff]+)',txt)
        if mn: name=mn.group(1)
    eps=pe=pb=None
    for p,v in [(r'EPS[^\d\-]*([+-]?[\d.]+)','e'),(r'(?:P/E|本益比)[^\d]*([+-]?[\d.]+)','p'),(r'(?:P/B|股價淨值)[^\d]*([+-]?[\d.]+)','b')]:
        m=re.search(p,txt)
        if m:
            if v=='e': eps=float(m.group(1))
            elif v=='p': pe=float(m.group(1))
            else: pb=float(m.group(1))
    return eps,pe,pb,name
def fetch_t86(code,date):
    d=get_json(f"https://www.twse.com.tw/fund/T86?response=json&date={date}&selectType=ALLBUT0999")
    if d.get('stat')=='OK':
        for r in d.get('data',[]):
            if str(r[0]).strip()==code:
                def p(s): return int(s.strip().replace(',','').replace('+','')) if s.strip() not in ('---','','--') else 0
                try: return p(r[4])//1000, p(r[10])//1000, p(r[18])//1000
                except: pass
    return 0,0,0
def fetch_tpex_t86(code,date):
    """上櫃三大法人（TPEX 新版 API）
    欄位順序：[0]代號 [1]名稱 [2-4]自營自行 [5-7]自營避險 [8-10]自營合計
    [11-13]投信 [14-16]外資不含自營 [17-19]外資自營 [20-22]外資合計 [23]三大合計
    值為「股」需 //1000 轉張"""
    y_ad=f"{date[:4]}/{date[4:6]}/{date[6:]}"
    d=get_json(f"https://www.tpex.org.tw/www/zh-tw/insti/dailyTrade?date={y_ad}&type=Daily")
    for t in d.get('tables',[]):
        for r in t.get('data',[]):
            if str(r[0]).strip()==code:
                def p(s): return int(str(s).strip().replace(',','')) if str(s).strip().replace(',','').lstrip('-').isdigit() else 0
                try: return p(r[22])//1000, p(r[13])//1000, p(r[23])//1000
                except: pass
    return 0,0,0

def detect_market(code):
    """偵測上市(twse)/上櫃(tpex) — 先嘗試 TWSE 最近月，失敗則為 TPEX"""
    ym=datetime.today().strftime('%Y%m')
    d=get_json(f"https://www.twse.com.tw/exchangeReport/STOCK_DAY?response=json&date={ym}01&stockNo={code}")
    if d.get('stat')=='OK' and d.get('data'):
        return 'twse'
    # 嘗試上個月（本月初可能還沒資料）
    prev=(datetime.today().replace(day=1)-timedelta(days=1)).strftime('%Y%m')
    d2=get_json(f"https://www.twse.com.tw/exchangeReport/STOCK_DAY?response=json&date={prev}01&stockNo={code}")
    return 'twse' if d2.get('stat')=='OK' and d2.get('data') else 'tpex'

# ── 主函式 ──
def run(CODE, NAME, TODAY=None,
        sector_name="電子零組件", catalyst="", catalyst_quality="中",
        sector_rs_5d=None, stock_rs_5d=None, margin_change=None, news_positive=True,
        market=None, product_mix="待補充"):

    if TODAY is None: TODAY=datetime.today().strftime('%Y%m%d')
    T0=time.time()

    # 自動偵測市場別
    if market is None:
        market = detect_market(CODE)
    is_otc = (market=='tpex')
    mkt_tag = "上櫃" if is_otc else "上市"
    print(f"[START] {CODE} {NAME}（{mkt_tag}）")

    months=[]
    d=datetime.strptime(TODAY,'%Y%m%d')
    for _ in range(7):
        months.append(d.strftime('%Y%m')); d=(d.replace(day=1)-timedelta(days=1))
    months=months[::-1]
    tdays=last_n_wd(5)
    margin_days=last_n_wd(6)  # 多抓一天算增減

    fetch_ohlcv_fn = fetch_tpex if is_otc else fetch_twse
    fetch_inst_fn  = fetch_tpex_t86 if is_otc else fetch_t86

    futs={}
    with ThreadPoolExecutor(max_workers=25) as ex:
        for ym in months: futs[ex.submit(fetch_ohlcv_fn,CODE,ym)]=('tw',ym)
        futs[ex.submit(fetch_five91,CODE)]=('f5',None)
        for td in tdays: futs[ex.submit(fetch_inst_fn,CODE,td)]=('t86',td)
        # 自動抓 TAIEX（近 2 個月，算 5 日漲幅）
        taiex_months=list(set([TODAY[:6], (datetime.strptime(TODAY,'%Y%m%d')-timedelta(days=35)).strftime('%Y%m')]))
        for ym in taiex_months: futs[ex.submit(fetch_taiex,ym)]=('idx',ym)
        # 自動抓融資（首尾兩天即可算增減）
        if margin_change is None:
            futs[ex.submit(fetch_margin,CODE,margin_days[0])]=('mgn','now')
            futs[ex.submit(fetch_margin,CODE,margin_days[-1])]=('mgn','old')

        tw_rows={}; f5r=(None,None,None,None); t86d={}
        idx_data=[]; mgn={'now':None,'old':None}
        for f in as_completed(futs):
            tag,m=futs[f]; res=f.result()
            if tag=='tw': tw_rows[m]=res
            elif tag=='f5': f5r=res
            elif tag=='t86': t86d[m]=res
            elif tag=='idx': idx_data.extend(res if res else [])
            elif tag=='mgn': mgn[m]=res

    tf=time.time()
    all_r=[r for ym in months for r in tw_rows.get(ym,[])]
    if is_otc:
        dt,O,H,L,C,V=parse_tpex_ohlcv(all_r)
    else:
        dt,O,H,L,C,V=parse_ohlcv(all_r)
    if len(C)<60: print("[ERR] 資料不足"); return

    dp=min(98,len(C))
    O=O[-dp:]; H=H[-dp:]; L=L[-dp:]; C=C[-dp:]; V=V[-dp:]; dt=dt[-dp:]
    x=np.arange(dp)

    ma5=sma_np(C,5); ma20=sma_np(C,20); ma60=sma_np(C,min(60,dp-1))
    rsi=rsi_np(C); Kd,Dd=kd_np(H,L,C); ml,sl2,hm=macd_np(C)
    bu,bm,bd=boll_np(C); atr=atr_np(H,L,C)

    lc=C[-1]; lr=rsi[-1] if not np.isnan(rsi[-1]) else 50
    lk=Kd[-1]; ld=Dd[-1]; lh=hm[-1]
    lbu=bu[-1] if not np.isnan(bu[-1]) else lc*1.05
    lbd=bd[-1] if not np.isnan(bd[-1]) else lc*0.95
    bb_p=0 if (lbu-lbd)==0 else (lc-lbd)/(lbu-lbd)*100
    v5=V[-5:].mean(); v20=V[-20:].mean(); vr=v5/v20 if v20>0 else 1
    rs10=(lc-C[-11])/C[-11]*100 if len(C)>=11 else 0
    # 停損停利：ATR(10) 動態，匹配 5-10 日持有期
    # 停損 = 1.5× ATR(10)，上限 8%（超過代表方向錯誤）
    # 停利 = 3× ATR(10)，無上限（讓獲利奔跑）
    # R/R = tp_pct / sl_pct（自然隨波動度變化，不再千篇一律）
    sl_pct=min(1.5*atr/lc*100, 8)
    tp_pct=3*atr/lc*100
    sl_p=round(lc*(1-sl_pct/100),1)
    tp_p=round(lc*(1+tp_pct/100),1)
    rr=round(tp_pct/sl_pct,2) if sl_pct>0 else 0

    eps,pe,pb,name_api=f5r
    if not NAME and name_api: NAME=name_api
    if eps is None: eps=8.0
    if pe is None: pe=round(lc/eps,1) if eps>0 else 18.
    if pb is None: pb=2.1

    f5_sum=0; t5_sum=0; tot_chip=0
    for td in tdays:
        fi,ti,ci=t86d.get(td,(0,0,0)); f5_sum+=fi; t5_sum+=ti; tot_chip+=ci

    # ── 自動計算參數（若未手動提供）──
    # stock_rs_5d：個股近 5 日漲幅 vs 大盤
    if stock_rs_5d is None:
        stock_5d = (C[-1]-C[-6])/C[-6]*100 if len(C)>=6 else 0
        if idx_data:
            idx_data.sort()
            idx_closes=[c for _,c in idx_data]
            if len(idx_closes)>=6:
                mkt_5d=(idx_closes[-1]-idx_closes[-6])/idx_closes[-6]*100
            else:
                mkt_5d=(idx_closes[-1]-idx_closes[0])/idx_closes[0]*100 if len(idx_closes)>=2 else 0
        else:
            mkt_5d=0
        stock_rs_5d = stock_5d - mkt_5d
        print(f"[AUTO]  stock_rs_5d={stock_rs_5d:+.2f}%（個股 {stock_5d:+.2f}% − 大盤 {mkt_5d:+.2f}%）")

    # sector_rs_5d：無法自動取產業指數，預設 = stock_rs_5d * 0.4（保守估計板塊約佔個股強度的 4 成）
    if sector_rs_5d is None:
        sector_rs_5d = stock_rs_5d * 0.4
        print(f"[AUTO]  sector_rs_5d={sector_rs_5d:+.2f}%（估算：個股 RS × 0.4）")

    # margin_change：融資 5 日增減
    if margin_change is None:
        if mgn['now'] is not None and mgn['old'] is not None:
            margin_change = mgn['now'] - mgn['old']
            print(f"[AUTO]  margin_change={margin_change:+,d} 張（{mgn['old']:,} → {mgn['now']:,}）")
        else:
            margin_change = 0
            print(f"[AUTO]  margin_change=0（API 無回傳）")

    # ── 評分 ──
    qm={'強':10,'中':7,'弱':3}
    s = {}
    s['eps']    = 5 if eps>0 else 0
    s['pe']     = 5 if pe<20 else (3 if pe<25 else 0)
    s['rev']    = 5 if news_positive else 2
    s['pb']     = 5 if pb<3 else 3
    s['fund']   = s['eps']+s['pe']+s['rev']+s['pb']

    s['cat_q']  = qm.get(catalyst_quality,7)
    s['cat_c']  = 6 if catalyst_quality=='強' else (4 if catalyst_quality=='中' else 1)
    s['cat_co'] = 4 if catalyst_quality=='強' else (2 if catalyst_quality=='中' else 1)
    s['sec_rs'] = 7 if sector_rs_5d>0 else (3 if sector_rs_5d>-1 else 0)
    s['stk_rs'] = 5 if stock_rs_5d>0 else 0
    s['etf_v']  = 4 if vr>1.2 else 2
    s['media']  = 4 if catalyst_quality in ('強','中') else 1
    s['sector'] = s['cat_q']+s['cat_c']+s['cat_co']+s['sec_rs']+s['stk_rs']+s['etf_v']+s['media']

    s['t_rs']   = 6 if rs10>0 else 0
    s['t_bb']   = 6 if 50<=bb_p<=100 else (3 if 30<=bb_p<50 else 0)
    s['t_vol']  = 5 if vr>1.0 else 0
    s['t_rsi']  = 3 if 45<=lr<=72 else 0
    s['tech']   = s['t_rs']+s['t_bb']+s['t_vol']+s['t_rsi']

    s['c_f5']   = 8 if f5_sum>0 else 0
    s['c_t5']   = 7 if t5_sum>0 else 0
    s['c_tot']  = 5 if tot_chip>0 else 0
    s['chip']   = s['c_f5']+s['c_t5']+s['c_tot']

    raw = s['fund']+s['sector']+s['tech']+s['chip']

    # ── 分級風險扣分 ──
    # 封頂型：直接限制最高分（致命風險，不該進場）
    # 扣分型：從原始分扣除（風險提示，但非致命）
    cap = None          # 封頂值（None = 不封頂）
    penalty = 0         # 累計扣分
    forced = []         # 風險標籤（顯示用）

    if eps <= 0:
        cap = 54; forced.append("虧損股 → 封頂 54")
    if s['sector'] < 20:
        cap = 54; forced.append("板塊動能不足 → 封頂 54")
    if lr > 75:
        penalty += 10; forced.append(f"RSI 超買（{lr:.0f}）→ -10")
    if f5_sum < -2000:
        penalty += 8; forced.append(f"外資 5 日賣超 {f5_sum:+,d} 張 → -8")
    if margin_change > 3000:
        penalty += 5; forced.append(f"融資暴增 +{margin_change:,} 張 → -5")

    # 計算最終分數：封頂優先，否則扣分
    if cap is not None:
        total = min(raw, cap)
    else:
        total = max(raw - penalty, 0)

    sig = "建議進場" if total>=75 else ("觀望" if total>=55 else "不建議")
    sig_c = GREEN if total>=75 else (GOLD if total>=55 else RED)
    sig_emoji = "🟢" if total>=75 else ("🟡" if total>=55 else "🔴")
    has_dg = (total != raw)

    dims = [
        ('基本面＆估值', s['fund'], 20, TEAL),
        ('產業動能 ⭐',  s['sector'], 40, GOLD),
        ('技術面',       s['tech'], 20, BLUE),
        ('籌碼面',       s['chip'], 20, PURPLE),
    ]

    tc=time.time()
    print(f"[CALC]  原始{raw}→{total}  {sig}  dims={[d[1] for d in dims]}")

    # ════════════════════════════════════════════════════════════════════════
    # PAGE 1
    # ════════════════════════════════════════════════════════════════════════
    fig1 = plt.figure(figsize=(22, 14.5), facecolor=BG)

    # 主佈局：左=圖表(4行)，右=一塊完整 axes
    gs = gridspec.GridSpec(5, 2,
        height_ratios=[0.28, 2.2, 0.60, 0.70, 0.70],
        width_ratios=[1.50, 1],
        hspace=0.40, wspace=0.20,
        left=0.04, right=0.97, top=0.96, bottom=0.04)

    # ── 標題列（雙行：大標 + 資訊行）──
    date_str = f"{TODAY[:4]}/{TODAY[4:6]}/{TODAY[6:]}"
    ax0 = fig1.add_subplot(gs[0, :]); ax0.set_facecolor(BG)
    ax0.set_xlim(0,1); ax0.set_ylim(0,1); ax0.axis('off')
    ax0.add_patch(FancyBboxPatch((0,0), 1, 1, boxstyle="round,pad=0.01",
        facecolor=WINE_DK, edgecolor=WINE, linewidth=1.8, transform=ax0.transAxes))
    ax0.text(0.50, 0.65, f"{CODE}  {NAME}　短線交易分析",
        color=GOLD, fontsize=16, fontweight='bold', va='center', ha='center')
    ax0.text(0.50, 0.28,
        f"{date_str}   ｜   {sector_name}   ｜   收盤 {lc:.1f}   ｜   EPS {eps:.1f}   PE {pe:.1f}x   PB {pb:.1f}x",
        color='#AAAAAA', fontsize=9, va='center', ha='center')

    # ── K 線 ──
    ax_k = fig1.add_subplot(gs[1, 0]); ax_k.set_facecolor(PANEL)
    for sp_ in ax_k.spines.values(): sp_.set_edgecolor(GRAY_DK)
    for i in range(dp):
        cl = WINE_LT if C[i]>=O[i] else GREEN
        ax_k.plot([i,i],[L[i],H[i]], color=cl, lw=0.8, zorder=2)
        bh = max(abs(C[i]-O[i]), 0.1); bb_ = min(C[i],O[i])
        ax_k.add_patch(mpatches.Rectangle((i-0.38, bb_), 0.76, bh,
            fc=cl, ec=cl, lw=0.5, zorder=3))
    for ma_, c_, lb_, lw_ in [(ma5,GOLD,'MA5',1.1),(ma20,BLUE,'MA20',1.3),(ma60,PURPLE,'MA60',1.1)]:
        xs = [i for i,v in enumerate(ma_) if not np.isnan(v)]
        if xs: ax_k.plot(xs, [ma_[i] for i in xs], color=c_, lw=lw_, label=lb_, zorder=4)
    xb = [i for i,v in enumerate(bu) if not np.isnan(v)]
    if xb:
        ax_k.plot(xb, [bu[i] for i in xb], color=GRAY, lw=0.7, ls='--', alpha=0.5, label='BB')
        ax_k.plot(xb, [bd[i] for i in xb], color=GRAY, lw=0.7, ls='--', alpha=0.5)
        ax_k.fill_between(xb, [bu[i] for i in xb], [bd[i] for i in xb], alpha=0.03, color=WHITE)
    ax_k.axhline(lc, color=GOLD, lw=0.7, ls=':', alpha=0.8)
    ax_k.axhline(sl_p, color=RED, lw=0.7, ls=':', alpha=0.8)
    ax_k.text(-0.5, lc, f'{lc:.0f}', color=GOLD, fontsize=8, va='center', ha='right')
    ax_k.text(-0.5, sl_p, f'{sl_p:.0f}', color=RED, fontsize=8, va='center', ha='right')
    pi = int(np.argmax(C))
    # 箭頭只在 K 線區域內標注，不超出邊界
    ann_x = min(pi+3, dp-1)
    ann_y = min(H[pi]+atr*0.8, H.max()+atr*0.5)
    ax_k.annotate(f'高 {C[pi]:.0f}', xy=(pi, H[pi]),
        xytext=(ann_x, ann_y),
        color=WHITE, fontsize=8, arrowprops=dict(arrowstyle='->', color=WHITE, lw=0.7))
    ax_k.set_xlim(-1, dp+5); ax_k.yaxis.tick_right()
    stp = 14; tkp = [i for i in range(0, dp, stp)]
    tkl = ['/'.join(dt[p].split('/')[1:]) for p in tkp]
    ax_k.set_xticks(tkp); ax_k.set_xticklabels(tkl, color=GRAY, fontsize=7)
    ax_k.tick_params(colors=GRAY, labelsize=8); ax_k.grid(axis='y', color=GRAY_DK, alpha=0.3)
    ax_k.legend(loc='upper left', fontsize=7, facecolor=PANEL, labelcolor=WHITE, framealpha=0.7)
    ax_k.set_title('K 線  ·  MA5 / MA20 / MA60  ·  布林通道', color=WHITE, fontsize=12, fontweight='bold', pad=8, loc='left')

    # ── 成交量 ──
    ax_v = fig1.add_subplot(gs[2, 0]); ax_v.set_facecolor(PANEL)
    for sp_ in ax_v.spines.values(): sp_.set_edgecolor(GRAY_DK)
    vc = [WINE_LT if C[i]>=O[i] else GREEN for i in range(dp)]
    ax_v.bar(x, V, color=vc, alpha=0.65, width=0.8)
    ax_v.axhline(v20, color=GOLD_DIM, lw=0.8, ls='--', alpha=0.7)
    ax_v.set_xlim(-1, dp+5); ax_v.set_xticks([])
    ax_v.yaxis.tick_right()
    ax_v.text(dp+0.5, v20, f'{v20:.0f}', color=GOLD_DIM, fontsize=7, va='center')
    ax_v.tick_params(colors=GRAY, labelsize=7); ax_v.grid(axis='y', color=GRAY_DK, alpha=0.3)
    vol_tag = "放量" if vr>1.3 else ("量增" if vr>1.0 else "縮量")
    if vr > 1.5:
        vol_cmt = "量能明顯放大，資金積極介入"
    elif vr > 1.0:
        vol_cmt = "量能溫和放大，買盤穩定進場"
    else:
        vol_cmt = "量能萎縮，觀望氣氛濃"
    ax_v.set_title(f'成交量（張）  ·  5 日均量 {v5:.0f}  /  20 日均量 {v20:.0f}  /  量比 {vr:.1f}  —  {vol_cmt}',
        color=WHITE, fontsize=10, pad=6, loc='left')

    # ── RSI ──
    ax_r = fig1.add_subplot(gs[3, 0]); ax_r.set_facecolor(PANEL)
    for sp_ in ax_r.spines.values(): sp_.set_edgecolor(GRAY_DK)
    xr = [i for i,v in enumerate(rsi) if not np.isnan(v)]
    if xr: ax_r.plot(xr, [rsi[i] for i in xr], color=BLUE, lw=1.2)
    for lvl, c_, ls_ in [(72,WINE_LT,'--'),(30,GREEN,'--'),(50,GRAY_DK,':')]:
        ax_r.axhline(lvl, color=c_, lw=0.6, ls=ls_, alpha=0.7)
    ax_r.set_ylim(0,100); ax_r.set_xlim(-1, dp+5); ax_r.set_xticks([])
    ax_r.yaxis.tick_right()
    rc = RED if lr>72 else (GREEN if lr<30 else WHITE)
    ax_r.text(dp+0.5, lr, f'{lr:.1f}', color=rc, fontsize=9, va='center', fontweight='bold')
    ax_r.tick_params(colors=GRAY, labelsize=7); ax_r.grid(axis='y', color=GRAY_DK, alpha=0.3)
    if lr > 75:
        rsi_cmt = "超買區，追高風險大，留意回檔"
    elif lr > 60:
        rsi_cmt = "偏強未超買，多方仍有動能"
    elif lr > 40:
        rsi_cmt = "中性，方向不明確"
    else:
        rsi_cmt = "偏弱" + ("，接近超賣" if lr<30 else "") + "，下跌動能趨緩"
    ax_r.set_title(f'RSI {lr:.0f}（{rsi_cmt}）  ·  KD  K={lk:.0f}  /  D={ld:.0f}',
        color=WHITE, fontsize=10, pad=6, loc='left')

    # ── MACD ──
    ax_m = fig1.add_subplot(gs[4, 0]); ax_m.set_facecolor(PANEL)
    for sp_ in ax_m.spines.values(): sp_.set_edgecolor(GRAY_DK)
    hc = [WINE_LT if v>=0 else GREEN for v in hm]
    ax_m.bar(x, hm, color=hc, alpha=0.55, width=0.8)
    xm2 = [i for i,v in enumerate(ml) if not np.isnan(v)]
    ax_m.plot(xm2, [ml[i] for i in xm2], color=GOLD, lw=1.1, label='MACD')
    ax_m.plot(xm2, [sl2[i] for i in xm2], color=WINE_LT, lw=0.9, ls='--', label='Signal')
    ax_m.axhline(0, color=GRAY_DK, lw=0.5)
    ax_m.set_xlim(-1, dp+5); ax_m.yaxis.tick_right()
    ax_m.legend(loc='upper left', fontsize=6, facecolor=PANEL, labelcolor=WHITE, framealpha=0.6)
    ax_m.set_xticks(tkp); ax_m.set_xticklabels(tkl, color=GRAY, fontsize=7)
    ax_m.text(dp+0.5, lh, f'{lh:.2f}', color=GOLD if lh>=0 else GREEN, fontsize=8, va='center')
    ax_m.tick_params(colors=GRAY, labelsize=7); ax_m.grid(axis='y', color=GRAY_DK, alpha=0.3)
    if lh > 0 and ml[-1] > sl2[-1]:
        macd_cmt = "快線在慢線上方，多方趨勢延續"
    elif lh > 0:
        macd_cmt = "柱正但縮小，多方動能轉弱"
    elif lh < 0 and ml[-1] < sl2[-1]:
        macd_cmt = "空方主導，短線偏空"
    else:
        macd_cmt = "空方力道減弱，觀察黃金交叉"
    ax_m.set_title(f'MACD（{lh:+.2f}）—  {macd_cmt}',
        color=WHITE, fontsize=10, pad=6, loc='left')

    # ───────────── 右側：單一 axes，手繪全部 ─────────────
    ax_R = fig1.add_subplot(gs[1:, 1]); ax_R.set_facecolor(BG)
    ax_R.set_xlim(0, 1); ax_R.set_ylim(0, 1); ax_R.axis('off')

    # ── 區塊 1：總分卡（y 0.86–0.99）──
    ax_R.add_patch(FancyBboxPatch((0.03, 0.86), 0.94, 0.13,
        boxstyle="round,pad=0.012", facecolor=PANEL, edgecolor=sig_c,
        linewidth=2.0, transform=ax_R.transAxes))
    # 分數居中
    ax_R.text(0.50, 0.945, f"{total}", color=sig_c, fontsize=42,
        fontweight='bold', va='center', ha='center', transform=ax_R.transAxes)
    ax_R.text(0.50, 0.885, f"{sig_emoji}  {sig}  ·  {total}/100",
        color=sig_c, fontsize=11, fontweight='bold',
        va='center', ha='center', transform=ax_R.transAxes)

    # ── 區塊 2：評分表格（y 0.20–0.84）──
    ax_R.add_patch(FancyBboxPatch((0.03, 0.20), 0.94, 0.64,
        boxstyle="round,pad=0.012", facecolor=PANEL, edgecolor=WINE_DK,
        linewidth=1.0, transform=ax_R.transAxes))

    ax_R.text(0.50, 0.815, '評分分佈', color=WHITE, fontsize=10,
        fontweight='bold', ha='center', va='center', transform=ax_R.transAxes)
    # 欄標題
    for cx_, clbl_ in [(0.08,'維度'),(0.60,'得分'),(0.74,'滿分'),(0.90,'達成')]:
        ax_R.text(cx_, 0.785, clbl_, color=GRAY, fontsize=7, va='center',
            transform=ax_R.transAxes)
    ax_R.plot([0.06, 0.96], [0.77, 0.77], color=GRAY_DK, lw=0.5,
        transform=ax_R.transAxes)

    # 四維度行
    row_h = 0.080
    y0 = 0.735
    for i, (lb, val, mx, cl) in enumerate(dims):
        y = y0 - i * row_h
        pct = val / mx * 100
        pct_c = GREEN if pct >= 80 else (GOLD if pct >= 50 else RED)
        ax_R.add_patch(Rectangle((0.05, y - 0.020), 0.006, 0.040,
            facecolor=cl, transform=ax_R.transAxes))
        ax_R.text(0.08, y, lb, color=cl, fontsize=9.5, fontweight='bold',
            va='center', transform=ax_R.transAxes)
        ax_R.text(0.62, y, f'{val}', color=WHITE, fontsize=11, fontweight='bold',
            va='center', ha='center', transform=ax_R.transAxes)
        ax_R.text(0.74, y, f'/{mx}', color=GRAY, fontsize=9,
            va='center', ha='center', transform=ax_R.transAxes)
        ax_R.text(0.90, y, f'{pct:.0f}%', color=pct_c, fontsize=9.5, fontweight='bold',
            va='center', ha='center', transform=ax_R.transAxes)
        if i < 3:
            ax_R.plot([0.06, 0.96], [y - 0.040, y - 0.040], color=GRAY_DK, lw=0.3,
                transform=ax_R.transAxes)

    # 小計線
    y_sub = y0 - 4 * row_h + 0.010
    ax_R.plot([0.06, 0.96], [y_sub + 0.020, y_sub + 0.020], color=WINE_DK, lw=0.8,
        transform=ax_R.transAxes)
    ax_R.text(0.08, y_sub, '小計', color=WHITE, fontsize=9.5, fontweight='bold',
        va='center', transform=ax_R.transAxes)
    ax_R.text(0.62, y_sub, f'{raw}', color=WHITE, fontsize=11, fontweight='bold',
        va='center', ha='center', transform=ax_R.transAxes)
    ax_R.text(0.74, y_sub, '/100', color=GRAY, fontsize=9,
        va='center', ha='center', transform=ax_R.transAxes)

    # 風險扣分行（若有）
    if has_dg:
        y_pen = y_sub - 0.060
        dg_label = '封頂調整' if cap is not None else '風險扣分'
        ax_R.add_patch(FancyBboxPatch((0.06, y_pen - 0.018), 0.88, 0.036,
            boxstyle="round,pad=0.003", facecolor=ORANGE, edgecolor='none',
            alpha=0.08, transform=ax_R.transAxes))
        ax_R.text(0.08, y_pen, f'⚠️ {dg_label}', color=ORANGE, fontsize=9,
            fontweight='bold', va='center', transform=ax_R.transAxes)
        ax_R.text(0.62, y_pen, f'-{raw - total}', color=ORANGE, fontsize=11,
            fontweight='bold', va='center', ha='center', transform=ax_R.transAxes)
        y_final = y_pen - 0.055
    else:
        y_final = y_sub - 0.055

    # 最終分數
    ax_R.plot([0.06, 0.96], [y_final + 0.022, y_final + 0.022],
        color=sig_c, lw=1.2, transform=ax_R.transAxes)
    ax_R.text(0.08, y_final, '最終', color=sig_c, fontsize=10,
        fontweight='bold', va='center', transform=ax_R.transAxes)
    ax_R.text(0.62, y_final, f'{total}', color=sig_c, fontsize=14,
        fontweight='bold', va='center', ha='center', transform=ax_R.transAxes)
    ax_R.text(0.74, y_final, '/100', color=GRAY, fontsize=9,
        va='center', ha='center', transform=ax_R.transAxes)
    ax_R.text(0.90, y_final, f'{sig_emoji} {sig}', color=sig_c, fontsize=9,
        fontweight='bold', va='center', ha='center', transform=ax_R.transAxes)

    # 風險原因（若有）
    if has_dg:
        y_reason = y_final - 0.040
        reason_txt = '、'.join(forced)
        ax_R.text(0.08, y_reason, reason_txt, color=ORANGE, fontsize=7,
            va='center', transform=ax_R.transAxes)

    # ── 區塊 3：交易參數卡片（y 0.01–0.18）──
    ax_R.add_patch(FancyBboxPatch((0.03, 0.01), 0.94, 0.17,
        boxstyle="round,pad=0.012", facecolor=PANEL, edgecolor=WINE_DK,
        linewidth=1.0, transform=ax_R.transAxes))
    ax_R.text(0.50, 0.160, '交易參數', color=WHITE, fontsize=9,
        fontweight='bold', ha='center', va='center', transform=ax_R.transAxes)

    cards = [
        ('停損', f'{sl_p}', f'-{sl_pct:.1f}%', RED),
        ('停利', f'{tp_p}', f'+{tp_pct:.1f}%', GREEN),
        ('R/R', f'{rr}:1', '✓ 達標' if rr>=2 else '偏低', GOLD),
    ]
    for ci, (lb, vl, sub, cl) in enumerate(cards):
        cx = 0.18 + ci * 0.26
        ax_R.add_patch(FancyBboxPatch((cx-0.09, 0.025), 0.20, 0.12,
            boxstyle="round,pad=0.006", facecolor=cl, alpha=0.06,
            edgecolor=cl, linewidth=0.7, transform=ax_R.transAxes))
        ax_R.text(cx+0.01, 0.130, lb, color=GRAY, fontsize=8,
            ha='center', va='center', transform=ax_R.transAxes)
        ax_R.text(cx+0.01, 0.085, vl, color=cl, fontsize=14,
            fontweight='bold', ha='center', va='center', transform=ax_R.transAxes)
        ax_R.text(cx+0.01, 0.042, sub, color=cl, fontsize=7.5,
            ha='center', va='center', transform=ax_R.transAxes, alpha=0.85)

    # ════════════════════════════════════════════════════════════════════════
    # PAGE 2
    # ════════════════════════════════════════════════════════════════════════
    fig2 = plt.figure(figsize=(22, 14.5), facecolor=BG)
    gs2 = gridspec.GridSpec(3, 2, height_ratios=[0.12, 1, 0.48],
        width_ratios=[0.45, 0.55], hspace=0.16, wspace=0.14,
        left=0.04, right=0.97, top=0.96, bottom=0.04)

    # 標題
    ax_t2 = fig2.add_subplot(gs2[0,:]); ax_t2.set_facecolor(BG)
    ax_t2.set_xlim(0,1); ax_t2.set_ylim(0,1); ax_t2.axis('off')
    ax_t2.add_patch(FancyBboxPatch((0,0),1,1, boxstyle="round,pad=0.008",
        facecolor=WINE_DK, edgecolor=WINE, linewidth=1.5, transform=ax_t2.transAxes))
    ax_t2.text(0.50, 0.65, f"{CODE}  {NAME}　評分明細 & 策略結論",
        color=GOLD, fontsize=16, fontweight='bold', va='center', ha='center')
    ax_t2.text(0.50, 0.28,
        f"{date_str}   ｜   {sector_name}   ｜   收盤 {lc:.1f}",
        color='#AAAAAA', fontsize=9, va='center', ha='center')

    # ── 左：評分表格 ──
    ax_tbl = fig2.add_subplot(gs2[1,0]); ax_tbl.set_facecolor(BG); ax_tbl.axis('off')
    ax_tbl.set_xlim(0,1); ax_tbl.set_ylim(0,1)
    ax_tbl.add_patch(FancyBboxPatch((0.01,0.01),0.98,0.98,
        boxstyle="round,pad=0.015", facecolor=PANEL, edgecolor=WINE_DK,
        linewidth=1.2, transform=ax_tbl.transAxes))

    tbl = [
        ('基本面＆估值',  s['fund'], 20, True,  TEAL,   ''),
        ('EPS（TTM）',    s['eps'],   5, False, WHITE,  f'{"✓" if eps>0 else "⚠️"} {eps:.1f} 元'),
        ('PE 本益比',     s['pe'],    5, False, WHITE,  f'{pe:.1f}x'),
        ('月營收 YoY',    s['rev'],   5, False, WHITE,  f'{"✓ 正向" if news_positive else "待確認"}'),
        ('PB 股價淨值',   s['pb'],    5, False, WHITE,  f'{pb:.1f}x'),
        ('SEP','','','','',''),
        ('產業動能 ⭐',    s['sector'],40, True,  GOLD,   ''),
        ('催化劑品質',     s['cat_q'], 10, False, WHITE,  f'{catalyst_quality}'),
        ('題材延續性',     s['cat_c'],  6, False, WHITE,  f'{"高" if s["cat_c"]>=5 else "中"}'),
        ('核心受益程度',   s['cat_co'], 4, False, WHITE,  ''),
        ('產業 vs 大盤',   s['sec_rs'], 7, False, WHITE,  f'{sector_rs_5d:+.1f}%'),
        ('個股 vs 產業',   s['stk_rs'], 5, False, WHITE,  f'{stock_rs_5d:+.1f}%'),
        ('ETF 量能',       s['etf_v'],  4, False, WHITE,  f'{"放大" if vr>1.2 else "持平"}'),
        ('媒體關注度',     s['media'],  4, False, WHITE,  ''),
        ('SEP','','','','',''),
        ('技術面',         s['tech'],  20, True,  BLUE,   ''),
        ('超額 RS10',      s['t_rs'],   6, False, WHITE,  f'{rs10:+.1f}%'),
        ('布林位置',       s['t_bb'],   6, False, WHITE,  f'{bb_p:.0f}%'),
        ('量能比 5/20d',   s['t_vol'],  5, False, WHITE,  f'{vr:.2f}'),
        ('RSI (14)',       s['t_rsi'],  3, False, WHITE,  f'{lr:.1f}'),
        ('SEP','','','','',''),
        ('籌碼面',         s['chip'],  20, True,  PURPLE, ''),
        ('外資 5 日',      s['c_f5'],   8, False, WHITE,  f'{f5_sum:+,d} 張'),
        ('投信 5 日',      s['c_t5'],   7, False, WHITE,  f'{t5_sum:+,d} 張'),
        ('三大法人合計',   s['c_tot'],  5, False, WHITE,  f'{tot_chip:+,d} 張'),
    ]
    rows_vis = [r for r in tbl if r[0] != 'SEP']
    n_rows = len(rows_vis)
    rh = 0.86 / n_rows
    ys = 0.94

    ax_tbl.text(0.04, 0.97, '評分明細', color=WHITE, fontsize=11,
        fontweight='bold', va='top', transform=ax_tbl.transAxes)
    for cx_, clbl_ in [(0.58,'得分'),(0.72,'滿分'),(0.88,'數據')]:
        ax_tbl.text(cx_, ys, clbl_, color=GRAY, fontsize=7,
            va='center', ha='center', transform=ax_tbl.transAxes)
    ys -= rh * 0.5

    for i, row in enumerate(rows_vis):
        lb, val, mx, hdr, cl, dtl = row
        y = ys - i * rh
        if i > 0:
            lc_ = WINE_DK if hdr else GRAY_DK
            lw_ = 0.8 if hdr else 0.3
            ax_tbl.plot([0.04,0.96],[y+rh*0.5, y+rh*0.5], color=lc_, lw=lw_,
                transform=ax_tbl.transAxes, zorder=1)
        fs = 9.5 if hdr else 8
        fw = 'bold' if hdr else 'normal'
        indent = 0.04 if hdr else 0.08
        if hdr:
            ax_tbl.add_patch(Rectangle((0.03, y-rh*0.35), 0.005, rh*0.7,
                facecolor=cl, transform=ax_tbl.transAxes))
        ax_tbl.text(indent, y, lb, color=cl if hdr else WHITE, fontsize=fs, fontweight=fw,
            va='center', transform=ax_tbl.transAxes)
        if isinstance(val, (int, float)):
            sc_c = cl if hdr else (GREEN if val==mx else WHITE)
            ax_tbl.text(0.58, y, f'{val}', color=sc_c, fontsize=fs, fontweight=fw,
                va='center', ha='center', transform=ax_tbl.transAxes)
            ax_tbl.text(0.67, y, f'/{mx}', color=GRAY, fontsize=7,
                va='center', ha='left', transform=ax_tbl.transAxes)
        if dtl:
            ax_tbl.text(0.78, y, dtl, color=GRAY, fontsize=7,
                va='center', ha='left', transform=ax_tbl.transAxes)

    # 合計
    yb = ys - n_rows * rh
    ax_tbl.plot([0.04,0.96],[yb+rh*0.4, yb+rh*0.4], color=WINE, lw=1.2,
        transform=ax_tbl.transAxes, zorder=1)
    ax_tbl.text(0.04, yb, '合計', color=WHITE, fontsize=10, fontweight='bold',
        va='center', transform=ax_tbl.transAxes)
    if has_dg:
        ax_tbl.text(0.50, yb, f'{raw}', color=ORANGE, fontsize=10,
            va='center', ha='right', transform=ax_tbl.transAxes)
        ax_tbl.text(0.53, yb, '→', color=GRAY, fontsize=9,
            va='center', ha='center', transform=ax_tbl.transAxes)
    ax_tbl.text(0.58, yb, f'{total}', color=sig_c, fontsize=14, fontweight='bold',
        va='center', ha='center', transform=ax_tbl.transAxes)
    ax_tbl.text(0.67, yb, '/100', color=GRAY, fontsize=9,
        va='center', ha='left', transform=ax_tbl.transAxes)
    ax_tbl.text(0.78, yb, f'{sig_emoji} {sig}', color=sig_c, fontsize=10,
        fontweight='bold', va='center', transform=ax_tbl.transAxes)

    # ── 右：四維度文字分析 ──
    ax_txt = fig2.add_subplot(gs2[1,1]); ax_txt.set_facecolor(BG); ax_txt.axis('off')
    ax_txt.set_xlim(0,1); ax_txt.set_ylim(0,1)

    pe_str = '合理' if pe<20 else ('偏高' if pe<28 else '過高')
    bb_str = '健康區' if 50<=bb_p<=90 else ('突破上軌 ⚠️' if bb_p>90 else '弱勢區')
    kd_str2 = '黃金交叉' if (lk>ld and lk<80) else ('超買' if lk>=80 else '待觀察')

    panels = [
        (TEAL, f'基本面＆估值  {s["fund"]}/20',
         f'EPS {eps:.1f} 元（{"正向" if eps>0 else "虧損"}）  PE {pe:.1f}x（{pe_str}）  PB {pb:.1f}x\n'
         f'▸ {"獲利穩健，估值合理，不構成進場障礙" if s["fund"]>=15 else "估值偏高或獲利疑慮，需更強題材支撐"}'),

        (GOLD, f'產業動能  {s["sector"]}/40  ⭐',
         f'催化劑：{catalyst[:40] if catalyst else sector_name+"題材"}（{catalyst_quality}）\n'
         f'產業 vs 大盤 {sector_rs_5d:+.1f}%  ·  個股 vs 產業 {stock_rs_5d:+.1f}%\n'
         f'▸ {"題材明確可延續，板塊資金流入，核心進場驅動力" if s["sector"]>=28 else "板塊動能待觀察，題材尚未充分發酵"}'),

        (BLUE, f'技術面  {s["tech"]}/20',
         f'RS10 {rs10:+.1f}%  布林 {bb_p:.0f}%（{bb_str}）  量比 {vr:.2f}  RSI {lr:.1f}\n'
         f'ATR {atr:.2f}  KD K={lk:.0f}/D={ld:.0f}（{kd_str2}）  MACD {lh:+.2f}\n'
         f'▸ {"技術強勢，量價配合" if s["tech"]>=15 else "技術面有超買或弱勢警示，進場需謹慎"}'),

        (PURPLE, f'籌碼面  {s["chip"]}/20',
         f'外資 5 日 {f5_sum:+,d} 張  投信 5 日 {t5_sum:+,d} 張  合計 {tot_chip:+,d} 張\n'
         + (f'⚠️ 融資暴增 +{margin_change:,} 張（散戶追高風險）\n' if margin_change>2000 else '')
         + f'▸ {"法人站買方，籌碼方向明確" if s["chip"]>=15 else "籌碼面偏弱，機構未明顯站買方"}'),
    ]

    box_h = 0.228; gap = 0.012; y_top = 0.97
    for i, (cl, title, body) in enumerate(panels):
        y = y_top - i * (box_h + gap)
        ax_txt.add_patch(FancyBboxPatch((0.01, y-box_h), 0.98, box_h,
            boxstyle="round,pad=0.010", facecolor=PANEL, edgecolor=GRAY_DK,
            linewidth=0.8, transform=ax_txt.transAxes))
        ax_txt.add_patch(Rectangle((0.01, y-box_h), 0.008, box_h,
            facecolor=cl, transform=ax_txt.transAxes))
        ax_txt.text(0.04, y-0.018, title, color=cl, fontsize=9.5, fontweight='bold',
            va='top', transform=ax_txt.transAxes)
        ax_txt.text(0.04, y-0.055, body, color=WHITE, fontsize=8, va='top',
            transform=ax_txt.transAxes, linespacing=1.55)

    # ── 底部：策略結論 ──
    ax_st = fig2.add_subplot(gs2[2,:]); ax_st.set_facecolor(BG); ax_st.axis('off')
    ax_st.set_xlim(0,1); ax_st.set_ylim(0,1)

    ax_st.add_patch(FancyBboxPatch((0.005,0.02),0.99,0.96,
        boxstyle="round,pad=0.012", facecolor=PANEL,
        edgecolor=sig_c, linewidth=2.5, transform=ax_st.transAxes))
    ax_st.add_patch(Rectangle((0.005,0.02),0.012,0.96,
        facecolor=sig_c, transform=ax_st.transAxes))

    if total>=75:
        action="建議進場"
        strat=f"四面向共振確認。以 {sl_p} 元停損（-{sl_pct:.0f}%），目標 {tp_p} 元（+{tp_pct:.0f}%），R/R = {rr}:1。"
    elif total>=55:
        action="觀望等待"
        wc=[]
        if lr>72: wc.append(f"RSI 回落至 65 以下（目前 {lr:.0f}）")
        if bb_p>90: wc.append("布林回歸上軌以內")
        if margin_change>2000: wc.append("融資去化")
        if not wc: wc.append("等待更明確進場訊號")
        strat=f"題材具潛力但時機未到。觀察：{' ／ '.join(wc)}"
    else:
        action="不建議進場"
        strat="多面向訊號不足，風險高於報酬，建議保持觀望。"

    ax_st.text(0.03, 0.88, '策略結論  STRATEGY', color=GRAY, fontsize=9,
        va='center', transform=ax_st.transAxes, fontstyle='italic')
    ax_st.text(0.22, 0.88, f'{sig_emoji}  {action}', color=sig_c, fontsize=15,
        fontweight='bold', va='center', transform=ax_st.transAxes)
    ax_st.plot([0.03,0.62],[0.78,0.78], color=GRAY_DK, lw=0.6, transform=ax_st.transAxes)
    ax_st.text(0.03, 0.65, strat, color=WHITE, fontsize=10.5, va='center',
        transform=ax_st.transAxes)

    if has_dg:
        ax_st.add_patch(FancyBboxPatch((0.025, 0.37), 0.60, 0.18,
            boxstyle="round,pad=0.005", facecolor=ORANGE, edgecolor='none',
            alpha=0.10, transform=ax_st.transAxes))
        dg_type = '封頂' if cap is not None else '風險扣分'
        ax_st.text(0.03, 0.50,
            f"⚠️ {dg_type}：{raw} → {total}（-{raw-total}）",
            color=ORANGE, fontsize=9, fontweight='bold', va='center',
            transform=ax_st.transAxes)
        # 列出具體原因
        ax_st.text(0.03, 0.42,
            '、'.join(forced),
            color=ORANGE, fontsize=8, va='center', transform=ax_st.transAxes)
        rule_y = 0.26
    else:
        rule_y = 0.42

    ax_st.text(0.03, rule_y, '評分標準：', color=GRAY, fontsize=8, va='center',
        transform=ax_st.transAxes)
    ax_st.text(0.11, rule_y, '≥75 🟢 建議進場', color=GREEN, fontsize=8, va='center',
        transform=ax_st.transAxes)
    ax_st.text(0.28, rule_y, '55–74 🟡 觀望', color=GOLD, fontsize=8, va='center',
        transform=ax_st.transAxes)
    ax_st.text(0.42, rule_y, '<55 🔴 不建議', color=RED, fontsize=8, va='center',
        transform=ax_st.transAxes)

    rr_y = rule_y - 0.12
    ax_st.text(0.03, rr_y,
        f'R/R = 停利幅度 ÷ 停損幅度 = {tp_pct:.1f}% ÷ {sl_pct:.1f}% = {rr}:1  '
        f'→ 每承擔 1 元風險預期獲利 {rr} 元' + ('  ✓' if rr>=2 else '（建議 ≥ 2:1）'),
        color=GRAY, fontsize=7.5, va='center', transform=ax_st.transAxes)

    for xi, (lb, vl, sub, cl2) in enumerate([
        ('停損', f'{sl_p}', f'-{sl_pct:.0f}%', RED),
        ('停利', f'{tp_p}', f'+{tp_pct:.0f}%', GREEN),
        ('R/R', f'{rr}:1', '✓' if rr>=2 else '偏低', GOLD),
    ]):
        xp = 0.72 + xi * 0.09
        ax_st.add_patch(FancyBboxPatch((xp-0.035, 0.10), 0.07, 0.80,
            boxstyle="round,pad=0.008", facecolor=cl2, alpha=0.08,
            edgecolor=cl2, linewidth=1.0, transform=ax_st.transAxes))
        ax_st.text(xp, 0.82, lb, color=GRAY, fontsize=8, ha='center', va='center',
            transform=ax_st.transAxes)
        ax_st.text(xp, 0.52, vl, color=cl2, fontsize=13, fontweight='bold',
            ha='center', va='center', transform=ax_st.transAxes)
        ax_st.text(xp, 0.25, sub, color=cl2, fontsize=8, ha='center', va='center',
            transform=ax_st.transAxes, alpha=0.8)

    # ── 儲存 ──
    outdir = os.path.expanduser("~/Desktop/Stock Analysis/Reports/SwingTrade")
    os.makedirs(outdir, exist_ok=True)
    path = os.path.join(outdir, f"{CODE}_{NAME}_短線分析_{TODAY}.pdf")
    with PdfPages(path) as pdf:
        pdf.savefig(fig1, dpi=150,  facecolor=BG)
        pdf.savefig(fig2, dpi=150,  facecolor=BG)
    plt.close('all')

    te = time.time()
    print(f"[PDF]   → {path}")
    print(f"[TIME]  抓取 {tf-T0:.1f}s  計算 {tc-tf:.2f}s  PDF {te-tc:.1f}s  ⏱ 總計 {te-T0:.1f}s")

    # ── Company Overview 自動 Append ──
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Alignment as XLAlign
        from pathlib import Path as P
        ov_folder = P.home() / 'Desktop/Stock Analysis'
        ov_files  = sorted(ov_folder.glob('Company Overview_*.xlsx'))
        if ov_files:
            md = get_json(f"https://five91.onrender.com/api/metrics?stock_id={CODE}")
            sd = md.get('stocks',{}).get(CODE,{})
            ov_wb = openpyxl.load_workbook(ov_files[-1])
            ov_ws = ov_wb['公司追蹤清單']
            no_fill = PatternFill(fill_type=None)
            ctr = XLAlign(horizontal='center', vertical='center', wrap_text=True)
            found = None
            for row in ov_ws.iter_rows(min_row=2):
                if str(row[1].value)==CODE: found=row[0].row; break
            if not found:
                ov_ws.append([None]*13); found=ov_ws.max_row
                seq = max((ov_ws.cell(r,1).value or 0) for r in range(2,found) if ov_ws.cell(r,1).value)
                ov_ws.cell(found,1).value = seq+1
            row_vals = [ov_ws.cell(found,1).value, CODE, NAME or sd.get('name',''),
                sd.get('revenue_ttm'), sd.get('gross_margin'), sd.get('operating_margin'),
                sd.get('net_margin'), sd.get('net_income_ttm'), sd.get('share_capital'),
                sd.get('eps'), product_mix, sd.get('category','電子零組件業'), catalyst[:50] if catalyst else sector_name]
            for ci,v in enumerate(row_vals,1):
                c=ov_ws.cell(found,ci); c.value=v; c.fill=no_fill; c.alignment=ctr
            from datetime import date as _date
            td=_date.today().strftime('%Y%m%d')
            new_ov = ov_folder / f'Company Overview_{td}.xlsx'
            ov_wb.save(new_ov)
            if ov_files[-1]!=new_ov and ov_files[-1].exists(): os.remove(ov_files[-1])
            print(f"[OV]    → Company Overview_{td}.xlsx  ({CODE} {NAME} upserted)")
    except Exception as e:
        print(f"[OV]    ⚠️ append 失敗：{e}")

    return total, sig, te-T0

if __name__=='__main__':
    # v9：只需 catalyst + quality，其餘自動抓；支援上市+上櫃
    run("6651", "全宇昕", "20260402",
        sector_name="功率半導體（分離式元件）",
        catalyst="瀚荃公開收購 5-15% 股權（每股 104 元），併購題材帶動股價",
        catalyst_quality="中")
