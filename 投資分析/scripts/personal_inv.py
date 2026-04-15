#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""個人投資回報分析 — 4頁 PDF 橫式 A4（依 20260327 格式）"""

import os, datetime
import openpyxl
import matplotlib
import matplotlib.pyplot as plt
from pypdf import PdfReader, PdfWriter

matplotlib.use('Agg')

# ── 路徑（每次執行自動帶入當日日期）────────────────────────────────
XL_PATH     = os.path.expanduser('~/Desktop/Personal Financial Management/Personal Financial Management.xlsx')
OUT_DIR     = os.path.expanduser('~/Desktop/Personal Financial Management/')
REPORT_DATE = datetime.datetime.now().strftime('%Y/%m/%d')   # ← 自動更新
today_str   = datetime.datetime.now().strftime('%Y%m%d')
OUT_PDF     = os.path.join(OUT_DIR, f'個人投資回報分析_{today_str}.pdf')

# ── 頁面尺寸（橫式 A4，對齊 20260327 模板：982.5×691.2 pts）────────
FIG_W, FIG_H = 13.65, 9.60

# ── 配色 ──────────────────────────────────────────────────────────
WINE    = '#7B1C2E'
WINE_DK = '#4A0E1A'
WINE_MD = '#9B2335'
GOLD    = '#B8860B'
GOLD_LT = '#D4A843'
GOLD_PAL= '#F5E6B8'
CREAM   = '#FDFBF7'
GRAY    = '#9E9E9E'
GRAY_LT = '#F0EDE8'
GRAY_DK = '#555555'
DIVIDER = '#C8B89A'
GREEN_C = '#2E6B30'
RED_C   = '#9B2335'
BLUE_C  = '#1565C0'
FONT    = 'Heiti TC'

plt.rcParams.update({
    'font.family': FONT,
    'font.size': 9,
    'axes.unicode_minus': False,
    'figure.dpi': 150,
})

def wine_bar(fig, y, h, color=None):
    """在 figure 座標繪製底色條（預設 WINE_DK，可傳入自訂色）"""
    import matplotlib.patches as mpatches
    rect = mpatches.Rectangle((0, y), 1, h,
        facecolor=color if color else WINE_DK, edgecolor='none',
        transform=fig.transFigure, zorder=0, clip_on=False)
    fig.add_artist(rect)

def text_ax(fig, rect):
    """透明文字層 axes（不蓋掉底色）"""
    ax = fig.add_axes(rect)
    ax.set_facecolor('none'); ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.axis('off')
    return ax

def colored_ax(fig, rect, color):
    """有底色的 axes：axis('off') 後強制顯示 patch"""
    ax = fig.add_axes(rect)
    ax.axis('off')
    ax.patch.set_facecolor(color)
    ax.patch.set_visible(True)
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    return ax

def bordered_ax(fig, rect, bg=CREAM, border=DIVIDER, lw=1.2):
    """有底色＋邊框的 axes（不呼叫 axis('off')，spine 保持可見）"""
    ax = fig.add_axes(rect, facecolor=bg)
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.tick_params(left=False, bottom=False, labelleft=False, labelbottom=False)
    for sp in ax.spines.values():
        sp.set_visible(True); sp.set_color(border); sp.set_linewidth(lw)
    return ax

# ── QR_ROWS（每月月底新增一行）───────────────────────────────────
QR_ROWS = {
    '2025/09': 48,
    '2025/10': 50,
    '2025/11': 52,
    '2025/12': 56,
    '2026/01': 60,
    '2026/02': 67,
    '2026/03': 71,
}
MONTHS       = list(QR_ROWS.keys())
SHORT_MONTHS = ['25/09','25/10','25/11','25/12','26/01','26/02','26/03']

# ── 讀取 Excel ────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XL_PATH, data_only=True)
ws = wb['Investment']
CUT = datetime.datetime(2025, 9, 1)

all_rows = []
for r in range(2, ws.max_row + 1):
    name  = ws.cell(r, 2).value
    buy_d = ws.cell(r, 5).value
    if not name or not isinstance(buy_d, datetime.datetime): continue
    all_rows.append({
        'row': r, 'name': name,
        'qty':    ws.cell(r, 3).value or 0,
        'buy_p':  ws.cell(r, 4).value or 0,
        'buy_d':  buy_d,
        'cost':   ws.cell(r, 6).value or 0,
        'sell_p': ws.cell(r, 7).value,
        'sell_d': ws.cell(r, 8).value,
        'days':   ws.cell(r, 10).value or 0,
        'ret':    ws.cell(r, 11).value or 0,
        'paper':  ws.cell(r, 13).value or 0,
        'profit': ws.cell(r, 16).value or 0,
    })

recent      = [r for r in all_rows if r['buy_d'] >= CUT]
completed_r = [r for r in recent if r['sell_d']]
open_pos_r  = [r for r in recent if not r['sell_d']]
tot_real    = sum(t['profit'] for t in completed_r)
tot_unreal  = sum(p['paper']  for p in open_pos_r)
tot_pnl     = tot_real + tot_unreal
trades      = completed_r + open_pos_r
wins        = sum(1 for t in trades
                  if (t['sell_d'] and t['profit'] > 0)
                  or (not t['sell_d'] and t['paper'] > 0))
win_rate    = wins / len(trades) if trades else 0
avg_days    = sum(t['days'] for t in completed_r) / len(completed_r) if completed_r else 0
best_t      = max(completed_r, key=lambda x: x['profit']) if completed_r else None
worst_t     = min(completed_r, key=lambda x: x['profit']) if completed_r else None

monthly_pnl = {ym: ws.cell(ri, 17).value or 0 for ym, ri in QR_ROWS.items()}
monthly_ret = {ym: ws.cell(ri, 18).value or 0 for ym, ri in QR_ROWS.items()}

def month_win_rate(ym):
    mt = [t for t in trades if t['buy_d'].strftime('%Y/%m') == ym]
    if not mt: return 0
    w = sum(1 for t in mt
            if (t['sell_d'] and t['profit'] > 0)
            or (not t['sell_d'] and t['paper'] > 0))
    return w / len(mt)

monthly_wr = {ym: month_win_rate(ym) for ym in MONTHS}

# Benchmark（cols N~T = 14~20）
bm_cols   = list(range(14, 21))
mkt_ret   = [ws.cell(104, c).value for c in bm_cols]
fund_ret  = [ws.cell(106, c).value for c in bm_cols]
ind_ret   = [ws.cell(108, c).value for c in bm_cols]
beat_fund = [ws.cell(110, c).value for c in bm_cols]
beat_mkt  = [ws.cell(111, c).value for c in bm_cols]
BM_LABELS = ['25/10','25/11','25/12','26/01','26/02','26/03','03/31']  # ← 最後標籤每月更新

def cum_ret(rets):
    vals = [0.0]; cur = 1.0
    for r in rets:
        if r is not None: cur *= (1 + r)
        vals.append((cur - 1) * 100)
    return vals

def safe_pct(lst):
    return [v * 100 if v is not None else 0 for v in lst]

cum_i = cum_ret([v or 0 for v in ind_ret])
cum_f = cum_ret([v or 0 for v in fund_ret])
cum_m = cum_ret([v or 0 for v in mkt_ret])


# ══════════════════════════════════════════════════════════════════
# Page 1 — 績效總覽（橫式 A4）
# 版面：標題(8%) → Hero(9%) → 圖區(40%) → KPI卡(18%) → 底欄(17%)
# ══════════════════════════════════════════════════════════════════
def make_page1():
    fig = plt.figure(figsize=(FIG_W, FIG_H), facecolor=CREAM)

    # ── 標題列（酒紅底色 + 透明文字層）
    wine_bar(fig, 0.920, 0.080)
    ax_t = text_ax(fig, [0, 0.920, 1, 0.080])
    ax_t.text(0.015, 0.50, '個人投資回報分析',
              color=GOLD_LT, fontsize=22, fontweight='bold',
              va='center', fontfamily=FONT)
    ax_t.text(0.985, 0.50,
              f'2025/09 ~ 2026/03  |  製表日期 {REPORT_DATE}',
              color=GOLD_LT, fontsize=9, va='center',
              ha='right', fontfamily=FONT)

    # ── Hero 三格（CREAM 底 + DIVIDER 邊框）
    for i, (label, val, col) in enumerate([
        ('已實現損益', f'{tot_real:,.0f}',   GREEN_C if tot_real  >= 0 else RED_C),
        ('未實現損益', f'{tot_unreal:,.0f}', GOLD    if tot_unreal >= 0 else RED_C),
        ('總損益',     f'{tot_pnl:,.0f}',    WINE    if tot_pnl    >= 0 else RED_C),
    ]):
        ax = bordered_ax(fig, [0.030 + i * 0.325, 0.818, 0.295, 0.092])
        ax.text(0.5, 0.76, label, ha='center', va='center',
                color=GRAY_DK, fontsize=9, fontfamily=FONT)
        ax.text(0.5, 0.32, val,   ha='center', va='center',
                color=col, fontsize=20, fontweight='bold', fontfamily=FONT)

    # ── 月度損益長條圖（左半；台灣慣例：上漲紅、下跌綠；數字永遠在上方）
    ax_bar = fig.add_axes([0.045, 0.415, 0.430, 0.375])
    ax_bar.set_facecolor(CREAM)
    pnl_vals = [monthly_pnl[ym] for ym in MONTHS]
    bars = ax_bar.bar(SHORT_MONTHS, pnl_vals,
                      color=[WINE if v >= 0 else GREEN_C for v in pnl_vals],
                      width=0.55, zorder=3)
    for bar, val in zip(bars, pnl_vals):
        y_pos = max(bar.get_height(), 0) + max(pnl_vals) * 0.012
        ax_bar.text(bar.get_x() + bar.get_width() / 2,
                    y_pos, f'{val:,.0f}', ha='center', va='bottom',
                    fontsize=6.5, color=GRAY_DK, fontfamily=FONT)
    ax_bar.set_title('月度批次總損益', fontsize=10, color=GRAY_DK,
                     fontfamily=FONT, pad=5)
    ax_bar.set_ylabel('月度損益額 (NTD)', fontsize=7.5,
                      fontfamily=FONT, color=GRAY_DK)
    ax_bar.tick_params(axis='x', labelsize=7.5)
    ax_bar.tick_params(axis='y', labelsize=7)
    ax_bar.spines[['top', 'right']].set_visible(False)
    ax_bar.grid(axis='y', alpha=0.25, zorder=0, linestyle='--')
    ax_bar.axhline(0, color=GRAY_DK, linewidth=0.6)

    # ── 分隔線
    ax_div = colored_ax(fig, [0.482, 0.415, 0.002, 0.375], DIVIDER)

    # ── 散佈圖（右半）
    ax_sc = fig.add_axes([0.492, 0.422, 0.488, 0.365])
    ax_sc.set_facecolor(CREAM)
    if completed_r:
        xs = [t['days'] for t in completed_r]
        ys = [t['ret'] * 100 for t in completed_r]
        ss = [max(25, abs(t['cost']) / 7000) for t in completed_r]
        ax_sc.scatter(xs, ys, s=ss, c=WINE, alpha=0.75, zorder=3, label='已賣出')
    if open_pos_r:
        xo = [t['days'] for t in open_pos_r]
        yo = [t['paper'] / t['cost'] * 100 if t['cost'] else 0 for t in open_pos_r]
        so = [max(25, abs(t['cost']) / 7000) for t in open_pos_r]
        ax_sc.scatter(xo, yo, s=so, c=GOLD, alpha=0.85, zorder=3, label='持倉中')
    ax_sc.axhline(0, color=GRAY_DK, linewidth=0.5, linestyle='--')
    ax_sc.set_title('持有天數 VS 報酬率', fontsize=10, color=GRAY_DK,
                    fontfamily=FONT, pad=5)
    ax_sc.set_xlabel('持有天數', fontsize=8, fontfamily=FONT)
    ax_sc.set_ylabel('報酬率 (%)', fontsize=8, fontfamily=FONT)
    ax_sc.tick_params(labelsize=7.5)
    ax_sc.spines[['top', 'right']].set_visible(False)
    ax_sc.grid(alpha=0.2, zorder=0, linestyle='--')
    if completed_r or open_pos_r:
        ax_sc.legend(fontsize=7, prop={'family': FONT},
                     markerscale=0.8, framealpha=0.8,
                     edgecolor=DIVIDER, loc='upper right')

    # ── KPI 卡片（5 張均分，GRAY_LT 底 + DIVIDER 邊框）
    best_mo  = max(monthly_pnl.items(), key=lambda x: x[1])
    worst_mo = min(monthly_pnl.items(), key=lambda x: x[1])
    def mo_short(ym):  # '2025/09' → '25/09'
        return ym[2:4] + '/' + ym[5:]
    kpis = [
        ('已出售筆數',  f'{len(completed_r)} 筆',   GRAY_DK),
        ('勝率',        f'{win_rate * 100:.1f}%',    GREEN_C if win_rate >= 0.5 else RED_C),
        ('平均持有天數', f'{avg_days:.0f} 天',        GRAY_DK),
        ('最佳單月',
         f'{mo_short(best_mo[0])}\n+{best_mo[1]:,.0f}',   RED_C),   # 台灣：獲利=紅
        ('最差單月',
         f'{mo_short(worst_mo[0])}\n{worst_mo[1]:,.0f}',  GREEN_C), # 台灣：虧損=綠
    ]
    for i, (label, val, col) in enumerate(kpis):
        ax_k = bordered_ax(fig, [0.020 + i * 0.194, 0.192, 0.178, 0.192],
                           bg=GRAY_LT, border=DIVIDER, lw=0.9)
        ax_k.text(0.5, 0.75, label, ha='center', va='center',
                  fontsize=8.5, color=GRAY_DK, fontfamily=FONT)
        ax_k.text(0.5, 0.36, val, ha='center', va='center',
                  fontsize=13, fontweight='bold', color=col,
                  fontfamily=FONT, linespacing=1.5)

    # ── 底欄分析文字（酒紅底色 + 透明文字層）
    wine_bar(fig, 0.000, 0.184)
    ax_txt = text_ax(fig, [0, 0, 1, 0.184])
    lines = [
        f'分析期間 2025/09 ~ 2026/03，共 {len(trades)} 筆交易'
        f'（已賣出 {len(completed_r)} / 持倉中 {len(open_pos_r)}）',
        f'總損益 {tot_pnl:,.0f} 元'
        f'（已實現 {tot_real:,.0f} + 未實現 {tot_unreal:,.0f}），勝率 {win_rate * 100:.1f}%',
        (f'最佳單筆：{best_t["name"]} +{best_t["profit"]:,.0f} 元'
         f'  |  最差單筆：{worst_t["name"]} {worst_t["profit"]:,.0f} 元'
         if best_t and worst_t else ''),
        f'平均持有 {avg_days:.0f} 天，月度損益呈正向趨勢',
    ]
    for i, line in enumerate(lines):
        ax_txt.text(0.018, 0.82 - i * 0.21, line,
                    ha='left', va='center',
                    color=GOLD_LT, fontsize=8.5, fontfamily=FONT)

    tmp = os.path.join(OUT_DIR, '_p1.pdf')
    fig.savefig(tmp, format='pdf', facecolor=CREAM)
    plt.close(fig)
    return tmp


# ══════════════════════════════════════════════════════════════════
# Page 2 — 月度交易分析（橫式 A4）
# 版面：標題(5%) → 雙軸圖(32%) → 7月卡片(58%)
# ══════════════════════════════════════════════════════════════════
def make_page2():
    fig = plt.figure(figsize=(FIG_W, FIG_H), facecolor=CREAM)

    # ── 雙層標題：深酒紅（月度交易分析）+ 中酒紅（月度損益 & 勝率）
    wine_bar(fig, 0.960, 0.040)                              # Layer 1 WINE_DK
    ax_t1 = text_ax(fig, [0, 0.960, 1, 0.040])
    ax_t1.text(0.015, 0.50, '月度交易分析',
               color=GOLD_LT, fontsize=15, fontweight='bold',
               va='center', fontfamily=FONT)
    wine_bar(fig, 0.928, 0.032, WINE_MD)                    # Layer 2 WINE_MD
    ax_t2 = text_ax(fig, [0, 0.928, 1, 0.032])
    ax_t2.text(0.015, 0.50, '月度損益 & 勝率',
               color='white', fontsize=10,
               va='center', fontfamily=FONT)

    # ── 月度損益 & 勝率雙軸圖（上移確保月份標籤可見）
    ax_bar = fig.add_axes([0.07, 0.505, 0.86, 0.405])
    ax_bar.set_facecolor(CREAM)
    pnl_vals = [monthly_pnl[ym] for ym in MONTHS]
    bars = ax_bar.bar(SHORT_MONTHS, pnl_vals,
                      color=[WINE if v >= 0 else GREEN_C for v in pnl_vals],
                      width=0.50, zorder=3)
    for bar, val in zip(bars, pnl_vals):
        y_pos = max(bar.get_height(), 0) + max(pnl_vals) * 0.015
        ax_bar.text(bar.get_x() + bar.get_width() / 2,
                    y_pos, f'{val:,.0f}', ha='center', va='bottom',
                    fontsize=8, color=GRAY_DK, fontfamily=FONT)
    ax_bar.set_ylabel('損益（NTD）', fontsize=9, fontfamily=FONT)
    ax_bar.tick_params(axis='x', labelsize=9)
    ax_bar.tick_params(axis='y', labelsize=8)
    ax_bar.spines[['top', 'right']].set_visible(False)
    ax_bar.grid(axis='y', alpha=0.2, zorder=0, linestyle='--')
    ax_bar.axhline(0, color=GRAY_DK, linewidth=0.5)
    ax_bar.set_title('', pad=2)   # 標題已移至雙層頁首

    ax2 = ax_bar.twinx()
    wr_vals = [monthly_wr[ym] * 100 for ym in MONTHS]
    ax2.plot(SHORT_MONTHS, wr_vals,
             color=GOLD, marker='D', linewidth=2, markersize=8, zorder=4)
    for i, (x, y) in enumerate(zip(SHORT_MONTHS, wr_vals)):
        ax2.text(i, y + 4, f'{y:.0f}%',
                 ha='center', va='bottom', fontsize=9,
                 color=GOLD, fontfamily=FONT, fontweight='bold')
    ax2.set_ylim(0, 105)
    ax2.set_ylabel('勝率 (%)', fontsize=9, fontfamily=FONT, color=GOLD)
    ax2.tick_params(axis='y', labelsize=8, colors=GOLD)
    ax2.spines[['top']].set_visible(False)

    # ── 7 月卡片（下半約 50%，整體往下）
    for i, ym in enumerate(MONTHS):
        l = 0.012 + i * 0.138
        ax_c = bordered_ax(fig, [l, 0.010, 0.130, 0.448],
                           bg=GRAY_LT, border=DIVIDER, lw=0.8)

        # wine 頂條
        ax_c.add_patch(plt.Rectangle((0, 0.910), 1, 0.090,
                                     facecolor=WINE_MD, edgecolor='none'))
        label2 = ('2025/' if SHORT_MONTHS[i].startswith('25') else '2026/') + SHORT_MONTHS[i][3:]
        ax_c.text(0.5, 0.955, label2, ha='center', va='center',
                  fontsize=9, fontweight='bold', color='white', fontfamily=FONT)

        pnl_v = monthly_pnl[ym]
        ret_v = monthly_ret[ym] * 100
        wr_v  = monthly_wr[ym] * 100
        pnl_c = GREEN_C if pnl_v >= 0 else RED_C

        ax_c.text(0.5, 0.845, f'{pnl_v:,.0f}',
                  ha='center', va='center', fontsize=10,
                  fontweight='bold', color=pnl_c, fontfamily=FONT)
        ax_c.text(0.5, 0.782, f'{ret_v:.1f}%',
                  ha='center', va='center', fontsize=8.5,
                  color=pnl_c, fontfamily=FONT)
        ax_c.text(0.5, 0.722, f'勝率 {wr_v:.0f}%',
                  ha='center', va='center', fontsize=7.5,
                  color=GRAY, fontfamily=FONT)
        ax_c.axhline(0.708, color=DIVIDER, linewidth=0.6,
                     xmin=0.05, xmax=0.95)

        # 個別交易（名稱左、損益右）
        mt = [t for t in trades if t['buy_d'].strftime('%Y/%m') == ym]
        wy = 0.672
        for t in mt:
            is_open  = not t['sell_d']
            pnl_disp = t['paper'] if is_open else t['profit']
            tc       = GREEN_C if pnl_disp >= 0 else RED_C
            sym      = '* ' if is_open else ''
            ax_c.text(0.05, wy, f"{sym}{t['name']}",
                      ha='left', va='center',
                      fontsize=6, color=GRAY_DK, fontfamily=FONT)
            ax_c.text(0.95, wy, f'{pnl_disp:,.0f}',
                      ha='right', va='center',
                      fontsize=6, color=tc, fontfamily=FONT)
            wy -= 0.065
            if wy < 0.03: break

    tmp = os.path.join(OUT_DIR, '_p2.pdf')
    fig.savefig(tmp, format='pdf', facecolor=CREAM)
    plt.close(fig)
    return tmp


# ══════════════════════════════════════════════════════════════════
# Page 3 — 交易明細與持倉（橫式 A4）
# 版面：標題(5%) → 全頁表格(95%)
# ══════════════════════════════════════════════════════════════════
def make_page3():
    fig = plt.figure(figsize=(FIG_W, FIG_H), facecolor=CREAM)

    # ── 雙層標題：深酒紅（交易明細與持倉）+ 中酒紅（全部交易明細）
    wine_bar(fig, 0.960, 0.040)                              # Layer 1 WINE_DK
    ax_t1 = text_ax(fig, [0, 0.960, 1, 0.040])
    ax_t1.text(0.015, 0.50, '交易明細與持倉',
               color=GOLD_LT, fontsize=15, fontweight='bold',
               va='center', fontfamily=FONT)
    wine_bar(fig, 0.928, 0.032, WINE_MD)                    # Layer 2 WINE_MD
    ax_t2 = text_ax(fig, [0, 0.928, 1, 0.032])
    ax_t2.text(0.015, 0.50, '全部交易明細',
               color='white', fontsize=10,
               va='center', fontfamily=FONT)

    # ── 表格（頂端調低，不與雙層標題重疊）
    ax_tbl = colored_ax(fig, [0.02, 0.020, 0.96, 0.900], CREAM)

    # 欄位 x 座標（移除舊標題，表頭從頂端開始）
    hx   = [0.005, 0.032, 0.190, 0.238, 0.308, 0.382, 0.440, 0.603, 0.772]
    hdrs = ['#', '股票名稱', '張數', '買入日', '賣出日', '天數', '成本', '損益', '報酬率']
    hy   = 0.970

    # 表頭（WINE_MD 底 white 字）
    ax_tbl.add_patch(plt.Rectangle((0, hy - 0.022), 1, 0.032,
                                   facecolor=WINE_MD, edgecolor='none'))
    for hxi, h in zip(hx, hdrs):
        ax_tbl.text(hxi + 0.004, hy - 0.005, h,
                    ha='left', va='center', fontsize=7.5,
                    fontweight='bold', color='white', fontfamily=FONT)

    # 資料列
    sorted_t = sorted(recent, key=lambda x: x['buy_d'])
    total    = len(sorted_t)
    row_h    = min(0.026, (hy - 0.055) / max(total, 1))
    fsize    = 6.0 if total > 30 else 7.0

    for j, t in enumerate(sorted_t):
        y = hy - 0.028 - j * row_h
        if y < 0.025: break
        ax_tbl.add_patch(plt.Rectangle(
            (0, y - row_h * 0.45), 1, row_h,
            facecolor=GRAY_LT if j % 2 == 0 else CREAM,
            edgecolor='none', zorder=1))

        is_open  = not t['sell_d']
        pnl_disp = t['paper'] if is_open else t['profit']
        ret_pct  = t['ret'] * 100
        pc       = RED_C if pnl_disp >= 0 else GREEN_C   # 台灣：獲利紅、虧損綠
        ret_str  = f'+{ret_pct:.2f}%' if ret_pct >= 0 else f'{ret_pct:.2f}%'
        sell_str = '持倉中' if is_open else t['sell_d'].strftime('%m/%d')

        vals  = [str(j + 1), t['name'], str(t['qty']),
                 t['buy_d'].strftime('%m/%d'), sell_str,
                 str(int(t['days'])), f"{t['cost']:,.0f}",
                 f'{pnl_disp:,.0f}', ret_str]
        vcols = [GRAY_DK] * 7 + [pc, pc]

        for hxi, v, vc in zip(hx, vals, vcols):
            ax_tbl.text(hxi + 0.004, y, v,
                        ha='left', va='center',
                        fontsize=fsize, color=vc,
                        fontfamily=FONT, zorder=2)

    ax_tbl.text(0.5, 0.008,
                f'共 {total} 筆交易（已賣出 {len(completed_r)} / 持倉中 {len(open_pos_r)}）',
                ha='center', va='bottom',
                fontsize=8, color=GRAY_DK, fontfamily=FONT)

    tmp = os.path.join(OUT_DIR, '_p3.pdf')
    fig.savefig(tmp, format='pdf', facecolor=CREAM)
    plt.close(fig)
    return tmp


# ══════════════════════════════════════════════════════════════════
# Page 4 — 績效比較（橫式 A4）
# 版面：標題(5%) → Grouped Bar(27%) → 累積圖(左57%) + 績效表(右57%)
# ══════════════════════════════════════════════════════════════════
def make_page4():
    fig = plt.figure(figsize=(FIG_W, FIG_H), facecolor=CREAM)

    # ── 標題
    wine_bar(fig, 0.950, 0.050)
    ax_t = text_ax(fig, [0, 0.950, 1, 0.050])
    ax_t.text(0.015, 0.50, '績效比較',
              color=GOLD_LT, fontsize=16, fontweight='bold',
              va='center', fontfamily=FONT)

    # ── Grouped Bar（上移，底部不顯示 x 標籤避免與下方重疊）
    n  = len(BM_LABELS); xs = list(range(n)); w = 0.24
    ax_gb = fig.add_axes([0.07, 0.665, 0.86, 0.250])
    ax_gb.set_facecolor(CREAM)
    ax_gb.bar([x - w for x in xs], safe_pct(ind_ret),  width=w, label='個人', color=WINE,   zorder=3)
    ax_gb.bar([x     for x in xs], safe_pct(fund_ret), width=w, label='基金', color=GOLD,   zorder=3)
    ax_gb.bar([x + w for x in xs], safe_pct(mkt_ret),  width=w, label='大盤', color=BLUE_C, zorder=3)
    ax_gb.set_xticks(xs); ax_gb.set_xticklabels([])   # 月份標籤在下方表格顯示，此處省略
    ax_gb.axhline(0, color=GRAY_DK, linewidth=0.5)
    ax_gb.set_ylabel('月度報酬率 (%)', fontsize=9, fontfamily=FONT)
    ax_gb.set_title('月度報酬率比較（個人 / 基金 / 大盤）',
                    fontsize=11, color=GRAY_DK, fontfamily=FONT, pad=6)
    ax_gb.legend(fontsize=9, prop={'family': FONT}, loc='upper right')
    ax_gb.spines[['top', 'right']].set_visible(False)
    ax_gb.grid(axis='y', alpha=0.2, linestyle='--', zorder=0)
    ax_gb.tick_params(axis='y', labelsize=8)

    # ── 累積報酬折線（左下，整體下移）
    ax_cum = fig.add_axes([0.05, 0.012, 0.545, 0.620])
    ax_cum.set_facecolor(CREAM)
    x_cum        = list(range(len(BM_LABELS) + 1))
    x_labels_cum = ['09(基)'] + BM_LABELS
    for cum_vals, label, color in [
        (cum_i, '個人', WINE),
        (cum_f, '基金', GOLD),
        (cum_m, '大盤', BLUE_C),
    ]:
        ax_cum.plot(x_cum, cum_vals,
                    marker='o', label=label, color=color,
                    linewidth=2, zorder=3)
        for xi, yv in zip(x_cum, cum_vals):
            off = 3 if label == '個人' else (-6 if label == '大盤' else 3)
            ax_cum.text(xi, yv + off, f'{yv:.1f}%',
                        ha='center', va='bottom',
                        fontsize=7, color=color, fontfamily=FONT)
    ax_cum.set_xticks(x_cum)
    ax_cum.set_xticklabels(x_labels_cum, fontsize=8)
    ax_cum.axhline(0, color=GRAY_DK, linewidth=0.5, linestyle='--')
    ax_cum.set_ylabel('累積報酬率 (%)', fontsize=9, fontfamily=FONT)
    ax_cum.set_title('累積報酬率', fontsize=11,
                     color=GRAY_DK, fontfamily=FONT, pad=6)
    ax_cum.legend(fontsize=8, prop={'family': FONT})
    ax_cum.spines[['top', 'right']].set_visible(False)
    ax_cum.grid(alpha=0.2, linestyle='--', zorder=0)
    ax_cum.tick_params(axis='y', labelsize=8)

    # ── 績效表（右下，整體下移，雙層表頭設計避免重疊）
    WINE_ROW = '#F9F0F2'
    ax_tbl = colored_ax(fig, [0.625, 0.012, 0.365, 0.620], CREAM)
    t_cols = ['月份', '個人', '基金', '贏基金', '贏大盤']
    t_cx   = [0.01, 0.22, 0.42, 0.62, 0.81]

    # 第一層：深酒紅標題帶「月度績效紀錄」
    ax_tbl.add_patch(plt.Rectangle((0, 0.930), 1, 0.055,
                                   facecolor=WINE_DK, edgecolor='none', zorder=1))
    ax_tbl.text(0.5, 0.958, '月度績效紀錄', ha='center', va='center',
                fontsize=9, fontweight='bold', color=GOLD_LT, fontfamily=FONT, zorder=2)

    # 第二層：中酒紅欄位名稱
    ax_tbl.add_patch(plt.Rectangle((0, 0.870), 1, 0.056,
                                   facecolor=WINE_MD, edgecolor='none', zorder=1))
    for cx, h in zip(t_cx, t_cols):
        ax_tbl.text(cx + 0.01, 0.898, h, ha='left', va='center',
                    fontsize=7.5, fontweight='bold',
                    color='white', fontfamily=FONT, zorder=2)

    # 資料列（白 / 極淡酒紅交替 + 細格線）
    row_gap = 0.100
    for ri in range(n):
        y  = 0.820 - ri * row_gap
        bg = WINE_ROW if ri % 2 == 0 else 'white'
        ax_tbl.add_patch(plt.Rectangle(
            (0, y - 0.040), 1, 0.082,
            facecolor=bg, edgecolor=DIVIDER, linewidth=0.4, zorder=1))

        iv   = safe_pct(ind_ret)[ri]
        fv   = safe_pct(fund_ret)[ri]
        bf   = beat_fund[ri]
        bm_v = beat_mkt[ri]
        ic   = RED_C if iv >= 0 else GREEN_C   # 台灣：獲利紅、虧損綠
        fc   = RED_C if fv >= 0 else GREEN_C
        bf_txt = '是' if bf   == 'Yes' else ('否' if bf   == 'No' else '—')
        bm_txt = '是' if bm_v == 'Yes' else ('否' if bm_v == 'No' else '—')
        bf_c   = RED_C if bf   == 'Yes' else GREEN_C   # 贏=紅、輸=綠
        bm_c   = RED_C if bm_v == 'Yes' else GREEN_C

        row_vals = [BM_LABELS[ri], f'{iv:.1f}%', f'{fv:.1f}%', bf_txt, bm_txt]
        row_cols = [GRAY_DK, ic, fc, bf_c, bm_c]
        for cx, v, vc in zip(t_cx, row_vals, row_cols):
            ax_tbl.text(cx + 0.01, y, v, ha='left', va='center',
                        fontsize=7.5, color=vc, fontfamily=FONT, zorder=2)

    ax_tbl.text(0.5, 0.025,
                f'累計：個人 {cum_i[-1]:.1f}%｜基金 {cum_f[-1]:.1f}%｜大盤 {cum_m[-1]:.1f}%',
                ha='center', va='bottom',
                fontsize=7.5, color=GRAY_DK, fontfamily=FONT)

    tmp = os.path.join(OUT_DIR, '_p4.pdf')
    fig.savefig(tmp, format='pdf', facecolor=CREAM)
    plt.close(fig)
    return tmp


# ══ 執行 & 合併 PDF ═══════════════════════════════════════════════
print('Page 1...'); p1 = make_page1()
print('Page 2...'); p2 = make_page2()
print('Page 3...'); p3 = make_page3()
print('Page 4...'); p4 = make_page4()
print('合併...')
writer = PdfWriter()
for tmp_pdf in [p1, p2, p3, p4]:
    writer.add_page(PdfReader(tmp_pdf).pages[0])
with open(OUT_PDF, 'wb') as f:
    writer.write(f)
for tmp_pdf in [p1, p2, p3, p4]:
    os.remove(tmp_pdf)
print(f'完成：{OUT_PDF}')
