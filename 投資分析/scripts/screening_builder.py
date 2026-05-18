#!/usr/bin/env python3
"""月初選股自動建表 v5
優化：D/E 讀前月工作表、H 用 TPEx+YF、K 產品佔比、缺失股本手動補
格式：完全複製上月工作表樣式
"""

import requests, openpyxl, time, concurrent.futures, os, shutil, warnings, copy
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import Counter
warnings.filterwarnings('ignore')

t0 = time.time()

# ══════════════════════════════════════════════════
#  CONFIG — 每月只改這區
# ══════════════════════════════════════════════════
SHEET       = "202604"
PREV_SHEET  = "202603"          # 上個月工作表名
D_COL_SRC   = 5                 # 上月 E 欄 (col index, 1-based)
E_COL_SRC   = 8                 # 上月 H 欄
D_LABEL     = "3/2收盤價"        # D 欄標題
E_LABEL     = "4/2收盤價"        # E 欄標題
H_LABEL     = "4/7收盤價"        # H 欄標題
H_DATE      = "20260407"        # 只有 H 需要 API 抓

STOCKS = [
    ('2360','致茂'),    ('8996','高力'),    ('3036','文曄'),    ('3131','弘塑'),
    ('2887','台新金'),  ('6187','萬潤'),    ('4971','IET-KY'),  ('2605','新興'),
    ('6683','雍智科技'),('6217','中探針'),  ('6640','均華'),    ('4768','晶呈科技'),
    ('7734','印能科技'),('8027','鈦昇'),    ('4904','遠傳'),    ('6584','南俊國際'),
    ('6291','沛亨'),    ('3026','禾伸堂'),  ('5386','青雲'),    ('7728','光焱科技'),
    ('6588','東典光電'),('6568','宏觀'),    ('6727','亞泰金屬'),('3010','華立'),
    ('4949','有成精密'),('4577','達航科技'),('3691','碩禾'),    ('2548','華固'),
    ('6715','嘉基'),    ('8438','昶昕'),    ('6679','鈺太'),    ('6861','睿生光電'),
    ('3209','全科'),    ('6894','衛司特'),  ('6693','廣閎科'),  ('8147','正淩'),
    ('1714','和桐'),    ('8091','翔名'),    ('6015','宏遠證'),  ('3555','博士旺')
]

# five91 缺失股 → 手動補 TTM EPS（股本改由官方 API 抓取）
MANUAL_EPS = {
    '2887': {'ttm_eps': 1.91, 'name': '台新金'},
    '8438': {'ttm_eps': 2.21, 'name': '昶昕'},
    '6894': {'ttm_eps': 10.99,'name': '衛司特'},
    '6015': {'ttm_eps': 1.12, 'name': '宏遠證'},
}

# K 欄：產品佔比
PRODUCT_MIX = {
    '2360': '半導體/光電測試 ~46%；ATE/EV量測 ~45%；統包方案 ~9%',
    '8996': '板式熱交換器(熱泵/空調) ~55%；燃料電池反應盒 ~25%；AI液冷散熱 ~20%',
    '3036': '資料中心/伺服器 ~36%；手機 ~19%；車用 ~13%；工業 ~12%；其他 ~20%',
    '3131': '濕製程設備 ~60%；配方化學品 ~20%；代理量測設備 ~15%',
    '2887': '個人金融(銀行/信用卡) ~56%；法人金融 ~30%；金融市場 ~14%',
    '6187': '半導體先進封裝設備(CoWoS/SoIC) ~95%；其他 ~5%',
    '4971': 'GaAs磊晶 ~30%；InP磊晶 ~33%；GaSb磊晶 ~25%；其他 ~12%',
    '2605': '散裝船(Capesize) ~70%；油輪(VLCC) ~29%；船舶管理 ~1%',
    '6683': 'IC測試載板(Socket/Carrier) ~100%',
    '6217': '探針卡/探針 ~78%；電子零件 ~17%；金屬加工 ~5%',
    '6640': '先進封裝設備(Chip Sorter+Die Bonder) ~85%；AOI/AI檢測 ~15%',
    '4768': '半導體精密化學品 ~60%；設備零組件/代理 ~30%；其他 ~10%',
    '7734': '先進封裝設備(壓力烤箱/Void Free) ~80%；老化測試機 ~10%；搬送系統 ~10%',
    '8027': '雷射設備 ~53%；SMD包裝材料 ~18%；自動化設備/零件 ~29%',
    '4904': '行動通訊 ~57%；固網/寬頻 ~18%；企業ICT ~18%；其他 ~7%',
    '6584': '鋼珠導軌(線性滑軌) ~96%；其他 ~4%',
    '6291': '電腦周邊(電聲/光纖) ~65%；網通(電源IC/廣告機) ~35%',
    '3026': '被動元件(MLCC) ~40%；主動元件 ~26%；系統模組 ~15%；其他 ~19%',
    '5386': 'AI伺服器代理 ~50%；記憶體模組 ~40%；電腦周邊 ~10%',
    '7728': '半導體光電測試設備 ~54%；模擬光源 ~35%；晶圓級光電偵測 ~5%',
    '6588': 'LED元件/模組 ~70%；光學電子零件 ~20%；其他 ~10%',
    '6568': '射頻IC(RF IC/調諧晶片) ~99%；委託設計 ~1%',
    '6727': '精密金屬沖壓件/模具/自動化設備 ~100%',
    '3010': '平面顯示器材料 ~29%；電子資通訊材料 ~29%；半導體材料 ~26%；其他 ~16%',
    '4949': '太陽能模組 ~60%；半導體設備零件 ~40%',
    '4577': '專業服務(閥門/管路維修) ~63%；設備販售 ~34%；其他 ~3%',
    '3691': '太陽能導電漿料(銀漿/TOPCon) ~90%；電池材料(LFP) ~6%；其他 ~4%',
    '2548': '住宅建案 ~70%；商辦/廠辦 ~25%；租金收入 ~5%',
    '6715': '高速電傳輸線 ~85%；光纖通訊傳輸線 ~8%；光學電子零件 ~6%；其他 ~1%',
    '8438': '再生金屬(銅鹽/PCB廢液回收) ~80%；特用化學品 ~11%；其他 ~9%',
    '6679': '數位MEMS麥克風IC ~75%；電源管理IC ~15%；類比麥克風 ~4%；其他 ~6%',
    '6861': '醫療X光平板感測器 ~70%；工業X光感測器 ~20%；醫療電子組裝 ~10%',
    '3209': '寬頻/無線設備通路(10G PON/FWA) ~80%；AI伺服器/光模組 ~15%；其他 ~5%',
    '6894': '再生金屬(銅廢液電解回收) ~58%；電解設備+耗材 ~42%',
    '6693': 'Power MOSFET ~60%；BLDC馬達驅動模組 ~25%；SoC風扇驅動IC ~15%',
    '8147': '高速高頻連接器 ~80%；背板/機箱零件 ~15%；模具 ~5%',
    '1714': '多元醇醚類/界面活性劑 ~50%；乙醇胺類 ~20%；其他石化原料 ~30%',
    '8091': '半導體設備耗材/材料 ~81%；設備零件代理 ~19%',
    '6015': '證券經紀 ~60%；自營交易 ~25%；承銷/財管 ~15%',
    '3555': '電子零件代理 ~60%；資訊系統整合 ~25%；生技 ~15%',
}

# L 欄：題材
NEWS_THEMES = {
    '2360': 'AI/EV測試',
    '8996': 'AI液冷+燃料電池',
    '3036': 'AI伺服器IC通路',
    '3131': 'CoWoS濕製程',
    '2887': '新光金合併',
    '6187': 'CoWoS封裝設備',
    '4971': '光通訊InP磊晶',
    '2605': '散裝航運地緣',
    '6683': 'AI晶片測試載板',
    '6217': 'AI探針卡轉單',
    '6640': 'CoWoS先進封裝',
    '4768': '半導體特化品/TGV',
    '7734': 'AI先進封裝製程',
    '8027': '雷射/FOPLP封裝',
    '4904': '5G+AI企業服務',
    '6584': 'AI伺服器滑軌',
    '6291': '電源IC/光纖',
    '3026': 'AI高壓MLCC',
    '5386': 'AI記憶體通路',
    '7728': 'CPO矽光子檢測',
    '6588': 'LED/光通訊',
    '6568': '光通訊/ASIC',
    '6727': 'AI資料中心精密件',
    '3010': 'AI半導體材料通路',
    '4949': '太陽能+半導體耗材',
    '4577': '半導體閥門控制',
    '3691': 'TOPCon銀漿+散熱',
    '2548': '都更豪宅新案',
    '6715': 'TB5/CPO高速線纜',
    '8438': 'PCB銅廢液循環',
    '6679': 'MEMS麥克風/車用',
    '6861': '醫療X光偵測器',
    '3209': 'FWA寬頻/AI光模組',
    '6894': '半導體廢液銅回收',
    '6693': 'AI伺服器MOSFET',
    '8147': 'AI水冷+高速連接器',
    '1714': '表面活性劑/陸市場',
    '8091': '2nm高階耗材',
    '6015': '證券經紀',
    '3555': '電子零件+生技',
}

# L 欄產業動能分群（3支以上同色標記）
THEME_GROUPS = {
    'AI伺服器':       {'color': 'FFCCCC', 'codes': ['3036','6683','6584','3026','5386','6727','6693','8147']},
    'CoWoS/先進封裝': {'color': 'E2CCFF', 'codes': ['3131','6187','6640','7734','8027']},
    '光通訊/CPO':     {'color': 'CCFFCC', 'codes': ['4971','7728','6588','6568','6715']},
    '半導體設備/耗材': {'color': 'CCE5FF', 'codes': ['2360','6217','4768','3010','4577','8091','4949','6894']},
}
# Build reverse lookup: code → color
THEME_COLOR = {}
for grp in THEME_GROUPS.values():
    for c in grp['codes']:
        THEME_COLOR[c] = grp['color']

SRC     = os.path.expanduser("~/Desktop/Stock Analysis/Jeff_Stock Analysis.xlsx")
DRAFT   = os.path.expanduser("~/Desktop/Stock Analysis/Jeff_Stock Analysis_Draft.xlsx")

# ══════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════
codes = [c for c,_ in STOCKS]

def yf_close(ticker, yyyymmdd):
    dt = datetime.strptime(yyyymmdd, "%Y%m%d")
    p1 = int(time.mktime(dt.replace(hour=0).timetuple()))
    p2 = int(time.mktime(dt.replace(hour=23, minute=59).timetuple()))
    try:
        r = requests.get(f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}",
            params={"period1":p1,"period2":p2,"interval":"1d"},
            headers={"User-Agent":"Mozilla/5.0"}, timeout=12)
        return round(r.json()["chart"]["result"][0]["indicators"]["quote"][0]["close"][0], 2)
    except: return None

def yahoo_taiex(yyyymmdd):
    dt = datetime.strptime(yyyymmdd, "%Y%m%d")
    p1 = int(time.mktime(dt.replace(hour=0).timetuple()))
    p2 = int(time.mktime(dt.replace(hour=23, minute=59).timetuple()))
    try:
        r = requests.get("https://query1.finance.yahoo.com/v8/finance/chart/%5ETWII",
            params={"period1":p1,"period2":p2,"interval":"1d"},
            headers={"User-Agent":"Mozilla/5.0"}, timeout=12)
        return round(r.json()["chart"]["result"][0]["indicators"]["quote"][0]["close"][-1], 2)
    except: return None

# ══════════════════════════════════════════════════
#  STEP 1: 讀前月工作表 → D/E 價格 + TAIEX
# ══════════════════════════════════════════════════
print("📖 [1/3] 讀取前月工作表 D/E 價格...")
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_prev = wb_src[PREV_SHEET]

# Build {code: row} index from prev sheet
prev_idx = {}
for row in range(2, 50):
    v = ws_prev.cell(row, 2).value
    if v: prev_idx[str(v)] = row

pd_map, pe_map = {}, {}
for code in codes:
    r = prev_idx.get(code)
    if r:
        pd_map[code] = ws_prev.cell(r, D_COL_SRC).value
        pe_map[code] = ws_prev.cell(r, E_COL_SRC).value

# TAIEX from prev sheet row 44
taiex_d = ws_prev.cell(44, D_COL_SRC).value
taiex_e = ws_prev.cell(44, E_COL_SRC).value
wb_src.close()

ok_D = sum(1 for v in pd_map.values() if v)
ok_E = sum(1 for v in pe_map.values() if v)
print(f"   ✅ D:{ok_D}/40  E:{ok_E}/40  TAIEX_D:{taiex_d}  TAIEX_E:{taiex_e}")

# ══════════════════════════════════════════════════
#  STEP 2: five91(財務) + TPEx batch(OTC H價) + YF(TWSE H價 + TAIEX)
# ══════════════════════════════════════════════════
print("📡 [2/3] five91 + TPEx + Yahoo Finance（並行）...")

def roc(d): return f"{int(d[:4])-1911}/{d[4:6]}/{d[6:8]}"

def fetch_shares_outstanding():
    """TWSE + TPEx 官方 API → {code: 已發行普通股數}"""
    shares = {}
    try:
        r = requests.get("https://openapi.twse.com.tw/v1/opendata/t187ap03_L",
            headers={"User-Agent":"Mozilla/5.0"}, timeout=20)
        for d in r.json():
            code = d.get("公司代號","").strip()
            s = d.get("已發行普通股數或TDR原股發行股數","").strip()
            if code and s:
                try: shares[code] = int(s)
                except: pass
    except: pass
    try:
        r = requests.get("https://www.tpex.org.tw/openapi/v1/mopsfin_t187ap03_O",
            headers={"User-Agent":"Mozilla/5.0"}, timeout=20)
        for d in r.json():
            code = d.get("SecuritiesCompanyCode","").strip()
            s = d.get("IssueShares","").strip()
            if code and s:
                try: shares[code] = int(s)
                except: pass
    except: pass
    return shares

def tpex_batch(yyyymmdd):
    """OTC: one batch call → {code: price}"""
    prices = {}
    try:
        r = requests.get(
            "https://www.tpex.org.tw/web/stock/aftertrading/daily_close_quotes/stk_quote_result.php",
            params={"l":"zh-tw","d":roc(yyyymmdd),"o":"json"},
            timeout=25, headers={"User-Agent":"Mozilla/5.0"})
        for table in r.json().get("tables",[]):
            for row in table.get("data",[]):
                try: prices[str(row[0]).strip()] = float(str(row[2]).replace(",",""))
                except: pass
    except: pass
    return prices

with concurrent.futures.ThreadPoolExecutor(max_workers=6) as ex:
    f_f91  = ex.submit(lambda: requests.get("https://five91.onrender.com/api/metrics",timeout=60).json().get("stocks",{}))
    f_tpH  = ex.submit(tpex_batch, H_DATE)
    f_taiH = ex.submit(yahoo_taiex, H_DATE)
    f_shr  = ex.submit(fetch_shares_outstanding)
    f91     = f_f91.result();  print(f"   ✅ five91: {len(f91)} 筆（財務指標）")
    tpex_h  = f_tpH.result();  print(f"   ✅ TPEx H: {len(tpex_h)} OTC")
    taiex_h = f_taiH.result(); print(f"   ✅ TAIEX_H: {taiex_h}")
    shares_map = f_shr.result(); print(f"   ✅ 已發行股數: {len(shares_map)} 筆（TWSE+TPEx）")

# H price: TPEx (OTC) → YF .TW (TWSE) → YF .TWO (fallback)
ph_map = {}
yf_h_needed = []
for code in codes:
    p = tpex_h.get(code)
    if p:
        ph_map[code] = p
    else:
        yf_h_needed.append(code)

def yf_h(code):
    p = yf_close(f"{code}.TW", H_DATE)
    if p is None: p = yf_close(f"{code}.TWO", H_DATE)
    return code, p

if yf_h_needed:
    print(f"   📡 YF H 補抓 TWSE {len(yf_h_needed)} 支...")
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
        for code, p in ex.map(yf_h, yf_h_needed):
            ph_map[code] = p

# D/E missing → YF for new stocks not in prev sheet
de_missing = [c for c in codes if pd_map.get(c) is None]
if de_missing:
    print(f"   📡 YF D/E 補抓新股 {len(de_missing)} 支...")
    d_month = D_LABEL.split("/")[0]
    d_day = D_LABEL.split("/")[1].replace("收盤價","")
    d_yyyymmdd = f"2026{int(d_month):02d}{int(d_day):02d}"
    e_month = E_LABEL.split("/")[0]
    e_day = E_LABEL.split("/")[1].replace("收盤價","")
    e_yyyymmdd = f"2026{int(e_month):02d}{int(e_day):02d}"

    def fetch_de_yf(code):
        pd = yf_close(f"{code}.TW", d_yyyymmdd) or yf_close(f"{code}.TWO", d_yyyymmdd)
        pe = yf_close(f"{code}.TW", e_yyyymmdd) or yf_close(f"{code}.TWO", e_yyyymmdd)
        return code, pd, pe
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as ex:
        for code, d, e in ex.map(fetch_de_yf, de_missing):
            if d: pd_map[code] = d
            if e: pe_map[code] = e

ok_H = sum(1 for v in ph_map.values() if v)
print(f"   ✅ H:{ok_H}/40")

# ══════════════════════════════════════════════════
#  STEP 3: 建 Excel — 完全複製上月格式
# ══════════════════════════════════════════════════
print(f"\n📊 [3/3] 建立工作表 {SHEET}（複製 {PREV_SHEET} 格式）...")
shutil.copy2(SRC, DRAFT)
wb = openpyxl.load_workbook(DRAFT)
ws_ref = wb[PREV_SHEET]

# Delete target sheet if exists
if SHEET in wb.sheetnames:
    del wb[SHEET]

# Create new sheet and copy column widths from ref
ws = wb.create_sheet(SHEET)

# Copy column widths
for col_letter in "ABCDEFGHIJKLMNOPQR":
    if col_letter in ws_ref.column_dimensions:
        ws.column_dimensions[col_letter].width = ws_ref.column_dimensions[col_letter].width

# Copy row heights
for r in [1, 44]:
    if r in ws_ref.row_dimensions:
        ws.row_dimensions[r].height = ws_ref.row_dimensions[r].height

# Grid lines off
ws.sheet_view.showGridLines = False

# ── HEADER ROW: copy style from ref, update D/E/H labels ──
label_override = {4: D_LABEL, 5: E_LABEL, 8: H_LABEL, 12: "題材", 15: "2025EPS", 17: "已發行普通股數"}
for col in range(1, 19):  # A-R
    ref_cell = ws_ref.cell(1, col)
    new_cell = ws.cell(1, col)
    # Copy value (override D/E/H labels)
    new_cell.value = label_override.get(col, ref_cell.value)
    # Copy style
    new_cell.font = copy.copy(ref_cell.font)
    new_cell.fill = copy.copy(ref_cell.fill)
    new_cell.alignment = copy.copy(ref_cell.alignment)
    new_cell.border = copy.copy(ref_cell.border)
    new_cell.number_format = ref_cell.number_format

# ── DATA ROWS ──
missing = []
for idx, (code, name_pdf) in enumerate(STOCKS):
    row = idx + 2

    # Find a ref row to copy style from (use row 2 as default template)
    ref_row = 2

    f = f91.get(code, {})
    manual = MANUAL_EPS.get(code, {})
    name    = f.get("name") or manual.get("name", name_pdf)
    ni_ttm  = f.get("net_income_ttm")
    sc      = f.get("share_capital")
    ttm_eps = round(ni_ttm/sc, 2) if ni_ttm is not None and sc and sc > 0 else manual.get("ttm_eps")
    shares  = shares_map.get(code)  # 已發行普通股數（官方 API）

    p_d = pd_map.get(code)
    p_e = pe_map.get(code)
    p_h = ph_map.get(code)
    if p_d is None and p_e is None:
        missing.append(code)

    # Auto-screen M column
    m_val = ""
    if ttm_eps is not None and ttm_eps < 0:
        m_val = "否"

    # Copy cell style from ref row, then set values
    for col in range(1, 19):
        ref_c = ws_ref.cell(ref_row, col)
        new_c = ws.cell(row, col)
        new_c.font = copy.copy(ref_c.font)
        # K(11) and M(13): no fill; M="否" gets gray later
        if col in (11, 13):
            new_c.fill = PatternFill()  # no fill
        else:
            new_c.fill = copy.copy(ref_c.fill)
        new_c.alignment = copy.copy(ref_c.alignment)
        new_c.border = copy.copy(ref_c.border)
        new_c.number_format = ref_c.number_format

    # A: 排序
    ws.cell(row, 1).value = idx + 1
    # B: 代號
    ws.cell(row, 2).value = int(code) if code.isdigit() else code
    # C: 名稱
    ws.cell(row, 3).value = name
    # D: price
    if p_d: ws.cell(row, 4).value = p_d
    # E: price
    if p_e: ws.cell(row, 5).value = p_e
    # F: 漲幅 (same formula style as 202603)
    ws.cell(row, 6).value = f"=E{row}/D{row}-1"
    # G: 排名
    ws.cell(row, 7).value = f"=RANK(F{row},$F$2:$F$41)"
    # H: price
    if p_h: ws.cell(row, 8).value = p_h
    # I: 漲幅
    ws.cell(row, 9).value = f"=H{row}/E{row}-1"
    # J: 排名
    ws.cell(row, 10).value = f"=RANK(I{row},$I$2:$I$41)"
    # K: 產業（產品佔比）— force black font (ref has white-on-colored)
    k_cell = ws.cell(row, 11)
    k_cell.value = PRODUCT_MIX.get(code, "")
    k_cell.font = Font(name="Calibri", size=12, color="000000")
    # L: 題材（產業動能分群上色）
    l_cell = ws.cell(row, 12)
    l_cell.value = NEWS_THEMES.get(code, "")
    tc = THEME_COLOR.get(code)
    if tc:
        l_cell.fill = PatternFill("solid", fgColor=tc)
    # M: 第一階段選擇
    m_cell = ws.cell(row, 13)
    if m_val:
        m_cell.value = m_val
        m_cell.fill = PatternFill("solid", fgColor="D9D9D9")  # gray for 否
        m_cell.font = Font(color="000000")  # black font
    else:
        m_cell.value = None
    # N: 第二階段選擇
    ws.cell(row, 14).value = None
    # O: EPS
    if ttm_eps is not None:
        ws.cell(row, 15).value = ttm_eps
    # P: Current PE — formula; show "EPS為負數" if EPS < 0
    if ttm_eps is not None and ttm_eps < 0:
        p_cell = ws.cell(row, 16)
        p_cell.value = "EPS為負數"
        p_cell.font = Font(color="C00000")  # red
    else:
        ws.cell(row, 16).value = f'=IF(AND(E{row}<>"",O{row}<>"",O{row}>0),ROUND(E{row}/O{row},1),"")'
    # Q: 張數
    if shares:
        ws.cell(row, 17).value = shares
    # R: 市值（億）— formula (已發行普通股數 × 收盤價 / 1億)
    ws.cell(row, 18).value = f'=IF(E{row}<>"",ROUND(Q{row}*E{row}/100000000,2),"")'

# ── Row 44: TAIEX (no label, match prev) ──
for col, val in [(4, taiex_d), (5, taiex_e), (8, taiex_h)]:
    if val:
        c44 = ws.cell(44, col)
        ref44 = ws_ref.cell(44, col)
        c44.value = val
        c44.font = copy.copy(ref44.font)
        c44.alignment = copy.copy(ref44.alignment)
        c44.number_format = ref44.number_format

# ── J46:L56 績效統計區 ──
# 計算重疊股（未跳出）
prev_codes = set()
for row in range(2, 50):
    v = ws_ref.cell(row, 2).value
    if v: prev_codes.add(str(v))
overlap = [c for c in codes if c in prev_codes]
overlap_set = set(overlap)
n_overlap = len(overlap)

# 重疊股 A~G 欄淺綠底色（F 欄 >20% 不覆蓋，由 CF 處理）
green_bg = PatternFill("solid", fgColor="C6EFCE")
for idx, (code, _) in enumerate(STOCKS):
    if code in overlap_set:
        row = idx + 2
        for col in range(1, 8):  # A~G
            if col == 6:  # F 欄：不覆蓋，讓 CF >20% 紅色優先
                continue
            ws.cell(row, col).fill = green_bg

stats = {
    46: ("個人績效", None),       # K46 留空（還沒選）
    47: ("基金績效", "=AVERAGE(I2:I41)"),
    48: ("加權指數", "=H44/E44-1"),
    49: ("打敗基金", '=IF(K46="","",IF(K46>K47,"Yes","No"))'),
    50: ("打敗大盤", '=IF(K46="","",IF(K46>K48,"Yes","No"))'),
    52: ("上漲家數", '=COUNTIF($I$2:$I$41,">0")'),
    53: ("下跌家數", '=COUNTIF($I$2:$I$41,"<0")'),
    55: ("未跳出", n_overlap),
    56: ("跳出", 40 - n_overlap),
}
l_formulas = {
    52: "=K52/(K52+K53)",
    53: "=K53/(K53+K52)",
}

for r, (label, val) in stats.items():
    # Copy style from ref
    for col in range(10, 13):
        ref_c = ws_ref.cell(r, col)
        new_c = ws.cell(r, col)
        new_c.font = copy.copy(ref_c.font)
        new_c.fill = copy.copy(ref_c.fill)
        new_c.alignment = copy.copy(ref_c.alignment)
        new_c.border = copy.copy(ref_c.border)
        new_c.number_format = ref_c.number_format
    ws.cell(r, 10).value = label
    if val is not None:
        ws.cell(r, 11).value = val
    if r in l_formulas:
        ws.cell(r, 12).value = l_formulas[r]

# ── Copy conditional formatting from ref sheet ──
for cf in ws_ref.conditional_formatting:
    for rule in cf.rules:
        ws.conditional_formatting.add(str(cf.sqref), rule)

# ── F column: >20% same style as I column CF (bgColor format) ──
from openpyxl.styles.differential import DifferentialStyle
f_dxf = DifferentialStyle(
    font=Font(color="FF9C0006"),
    fill=PatternFill(bgColor="FFFFC7CE")
)
from openpyxl.formatting.rule import Rule
ws.conditional_formatting.add(
    "F2:F41",
    Rule(type="cellIs", operator="greaterThan", formula=["0.2"], dxf=f_dxf)
)

wb.save(DRAFT)

elapsed = time.time() - t0
print(f"\n{'━'*52}")
print(f"✅ 完成！→ {DRAFT}")
print(f"   成功: {40-len(missing)}/40" + (f"  ⚠️ 缺失: {', '.join(missing)}" if missing else ""))
print(f"⏱️  總耗時: {elapsed:.1f} 秒")
print(f"{'━'*52}")
print(f"\n📌 待人工填寫: K欄（產業+產品佔比）、L欄（備註/題材）、M欄確認、N欄第二階段")
