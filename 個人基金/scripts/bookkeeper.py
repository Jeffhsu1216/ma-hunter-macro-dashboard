#!/usr/bin/env python3
"""
bookkeeper.py — 客戶投資組合自動記帳腳本
================================================
用法（CLI）:
  python3 bookkeeper.py <客戶名> '<trades_json>'

用法（import）:
  from bookkeeper import run
  run("林峻毅", trades)

trades 格式（list of dict）:
  買入: {"name":"昶昕","code":"8438","type":"buy","shares":6,"price":80.5,"fee":192,"date":"2026-04-08"}
  賣出: {"name":"昶昕","code":"8438","type":"sell","shares":6,"price":90.0,"fee":200,"tax":1440,"date":"2026-04-08"}

  shares 為張數（可為小數，如 0.5 表示零股 500 股）
"""

import sys, os, json, copy, glob, datetime, urllib.request
import openpyxl

CLIENTS_DIR = os.path.expanduser("~/Desktop/Clients/Excel")


# ─── Yahoo Finance 收盤價 ─────────────────────────────────────────────────────

def yf_close(code):
    """嘗試 .TW（上市）再 .TWO（上櫃），取最近有效收盤價"""
    for suffix in ['.TW', '.TWO']:
        try:
            url = (f"https://query1.finance.yahoo.com/v8/finance/chart/"
                   f"{code}{suffix}?range=5d&interval=1d")
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            data = json.loads(urllib.request.urlopen(req, timeout=10).read())
            closes = data['chart']['result'][0]['indicators']['quote'][0]['close']
            for c in reversed(closes):
                if c is not None:
                    return round(c, 2)
        except Exception:
            continue
    return None


# ─── Excel 檔案定位 ───────────────────────────────────────────────────────────

def find_latest_excel(client_name):
    """找 ~/Desktop/Clients/Excel/{客戶名}_YYYYMMDD.xlsx 中最新的檔案"""
    import re
    pattern = os.path.join(CLIENTS_DIR, f"{client_name}_*.xlsx")
    files = [f for f in glob.glob(pattern)
             if re.search(r'_\d{8}\.xlsx$', f)]
    if not files:
        raise FileNotFoundError(f"找不到客戶 {client_name} 的 Excel 檔案（{pattern}）")
    return sorted(files)[-1]


# ─── 工作表結構輔助 ───────────────────────────────────────────────────────────

def find_total_row_buy(ws):
    """找買入側「總計」列（B 欄 = '總計'）"""
    for r in range(4, 60):
        if ws.cell(r, 2).value == '總計':
            return r
    return 20  # 預設

def find_first_empty_buy_row(ws, total_row):
    """找買入側第一個空白行（B、C 欄均空白）"""
    for r in range(4, total_row):
        if ws.cell(r, 2).value is None and ws.cell(r, 3).value is None:
            return r
    return None

def find_last_buy_data_row(ws, total_row):
    """找買入側最後一筆有資料的列（B、C 欄均有值）"""
    last = None
    for r in range(4, total_row):
        if ws.cell(r, 2).value is not None and ws.cell(r, 3).value is not None:
            last = r
    return last

def eval_formula_value(raw):
    """簡易公式求值（支援加減整數，如 =982745+1667268）"""
    if isinstance(raw, (int, float)):
        return raw
    if isinstance(raw, str) and raw.startswith('='):
        try:
            return eval(raw[1:])
        except Exception:
            pass
    return raw or 0


# ─── 格式複製 ─────────────────────────────────────────────────────────────────

def copy_row_style(ws, src_row, dst_row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst.font      = copy.copy(src.font)
            dst.fill      = copy.copy(src.fill)
            dst.alignment = copy.copy(src.alignment)
            dst.border    = copy.copy(src.border)
            dst.number_format = src.number_format


# ─── SUM 公式更新 ─────────────────────────────────────────────────────────────

def update_buy_totals(ws, total_row, last_data_row):
    """更新買入側總計列 SUM 公式，擴展至 last_data_row"""
    r = total_row
    ws.cell(r, 7).value  = f"=SUM(G4:G{last_data_row})"   # G 買入成本
    ws.cell(r, 8).value  = f"=SUM(H4:H{last_data_row})"   # H 手續費
    ws.cell(r, 10).value = f"=SUM(J4:J{last_data_row})"   # J 損益
    ws.cell(r, 11).value = f"=J{r}/G{r}"                   # K 收益率


# ─── 主函式 ──────────────────────────────────────────────────────────────────

def run(client_name, trades):
    """
    client_name : str，客戶名稱（對應 ~/Desktop/Clients/{客戶名}_*.xlsx）
    trades      : list of dict，交易清單（格式見檔頭說明）
    回傳 dict   : {saved_to, old_cash, new_cash, trades_added}
    """
    excel_path = find_latest_excel(client_name)
    print(f"📂 開啟: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Investment Portfolio']

    # ── 讀取現金餘額 ──
    c25_raw  = ws.cell(25, 3).value
    old_cash = eval_formula_value(c25_raw)
    new_cash = old_cash
    print(f"💰 C25 現金餘額（期初）: {old_cash:,.0f}")

    # ── 今日日期 ──
    today     = datetime.date.today()
    today_str = f"{today.month}/{today.day}"  # 例: 4/8

    # ── 買入總計列位置 ──
    total_row_buy = find_total_row_buy(ws)
    print(f"📊 買入總計列: Row {total_row_buy}")

    # ── 更新 I3 收盤價標題 ──
    ws.cell(3, 9).value = f"{today_str}收盤價"

    inserted_rows = []

    for trade in trades:
        t_date  = datetime.datetime.strptime(trade['date'], '%Y-%m-%d')
        name    = trade['name']
        code    = str(trade['code'])
        t_type  = trade.get('type', 'buy')
        shares  = trade['shares']   # 張數，可為小數
        price   = trade['price']
        fee     = trade['fee']

        # ── 買入 ──────────────────────────────────────────────────────────────
        if t_type == 'buy':
            r = find_first_empty_buy_row(ws, total_row_buy)
            if r is None:
                print(f"⚠️  買入側已滿，無法新增 {name}")
                continue

            # 格式參考列（取已有資料的前一行；若都空白則用後一行）
            ref = find_last_buy_data_row(ws, total_row_buy)
            if ref:
                copy_row_style(ws, ref, r, 2, 11)

            # 寫入各欄
            ws.cell(r, 2).value  = t_date                           # B 購買日期
            ws.cell(r, 3).value  = name                             # C 股票名稱
            ws.cell(r, 4).value  = code                             # D 股票代號
            ws.cell(r, 5).value  = shares                           # E 買入張數
            ws.cell(r, 6).value  = price                            # F 買入價格
            ws.cell(r, 7).value  = f"=F{r}*E{r}*1000"              # G 買入成本
            ws.cell(r, 8).value  = fee                              # H 購買手續費
            ws.cell(r, 10).value = f"=(I{r}-F{r})*E{r}*1000"       # J 當前損益
            ws.cell(r, 11).value = f"=(I{r}-F{r})/F{r}"            # K 當前收益率

            # 日期格式
            ws.cell(r, 2).number_format = 'yyyy/mm/dd'

            # 抓收盤價 → I 欄
            close_px = yf_close(code)
            if close_px:
                ws.cell(r, 9).value = close_px
                print(f"  ✅ {name}({code}) Row {r}  {shares}張@{price}  收盤:{close_px}")
            else:
                print(f"  ⚠️  {name}({code}) Row {r}  收盤價抓取失敗，I{r} 留空")

            # 現金扣減
            cost = price * shares * 1000
            new_cash -= (cost + fee)

            inserted_rows.append(r)

        # ── 賣出（基本邏輯，M~Y 欄詳細對應待未來優化）────────────────────────
        elif t_type == 'sell':
            tax  = trade.get('tax', 0)
            sell_amount = price * shares * 1000
            new_cash += (sell_amount - fee - tax)
            print(f"  💸 {name} 賣出 {shares}張@{price}  入帳:{sell_amount-fee-tax:,.0f}  現金:{new_cash:,.0f}")
            # TODO: 寫入賣出側 M~Y 欄（待下次賣出實測後優化）

    # ── 更新 SUM 公式 ──
    if inserted_rows:
        last_used = max(inserted_rows)
        update_buy_totals(ws, total_row_buy, last_used)
        print(f"\n📐 SUM 公式更新至 Row {last_used}")

    # ── 寫入 C25 現金餘額（數值，不用公式）──
    ws.cell(25, 3).value = round(new_cash)
    print(f"💰 C25 更新: {old_cash:,.0f} → {round(new_cash):,.0f}")

    # ── 儲存（更新檔名日期）──
    new_filename = f"{client_name}_{today.strftime('%Y%m%d')}.xlsx"
    new_path     = os.path.join(CLIENTS_DIR, new_filename)
    wb.save(new_path)
    print(f"\n✅ 已儲存: {new_path}")

    return {
        "saved_to"    : new_path,
        "old_cash"    : int(old_cash),
        "new_cash"    : round(new_cash),
        "trades_added": len(inserted_rows),
    }


# ─── CLI 入口 ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    client_name = sys.argv[1]
    trades      = json.loads(sys.argv[2])
    result      = run(client_name, trades)
    print(json.dumps(result, ensure_ascii=False, indent=2))
