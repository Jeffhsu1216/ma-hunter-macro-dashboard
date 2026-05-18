#!/usr/bin/env python3
"""全客戶 AUM 彙總 PDF — 與 2026/04/01 定稿版一致"""
import os, datetime
from pathlib import Path
from openpyxl import load_workbook
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import numpy as np

# ── 配色 ──
BG      = '#1a0a0a'
WINE    = '#8B1A1A'
GOLD    = '#C8A951'
LGOLD   = '#D4BC7A'
WARM    = '#F5F0E8'
CARD_BG = '#2A1515'
BORDER  = '#4A2525'
RED     = '#E74C3C'
GREEN   = '#4CAF50'
GRAY    = '#9E8E7E'
PIE_C   = ['#8B1A1A','#C8A951','#6A5ACD','#2E8B8B','#5B8B3A','#C05A20']

heiti = fm.FontProperties(fname='/System/Library/Fonts/STHeiti Medium.ttc')
heiti_b = fm.FontProperties(fname='/System/Library/Fonts/STHeiti Medium.ttc', weight='bold')

def color_val(v):
    if v is None: return GOLD
    return RED if v >= 0 else GREEN

def run(TODAY=None):
    today = TODAY or datetime.date.today()
    today_str = today.strftime('%Y/%m/%d')
    today_fn  = today.strftime('%Y%m%d')

    clients_dir = Path.home() / 'Desktop/Clients/Excel'
    out_dir = Path.home() / 'Desktop/Clients/AUM_Overview'
    out_dir.mkdir(exist_ok=True)

    # ── 客戶清單 ──
    clients = [
        {'name':'小阿姨', 'type':'Non-Fee',   'prefix':'小阿姨_', 'init':1318502},
        {'name':'林峻毅', 'type':'Fee-Based', 'prefix':'林峻毅_', 'init':2000000},
        {'name':'靜怡',   'type':'Fee-Based', 'prefix':'靜怡_',   'init':600000},
        {'name':'自營',   'type':'自營',       'prefix':'Jeff_',   'init':4233490},
        {'name':'Gary',   'type':'Fee-Based', 'prefix':'Gary_',   'init':4000000},
        {'name':'老哥',   'type':'Non-Fee',   'prefix':'Josh_',   'init':None},
    ]

    # ── 讀取 Excel ──
    data = []
    excel_dates = []
    for c in clients:
        import re
        files = sorted(f for f in clients_dir.glob(f"{c['prefix']}*.xlsx")
                       if re.search(r'_\d{8}\.xlsx$', f.name))
        if not files:
            data.append({**c, 'aum':0, 'i25':0, 'pnl':0, 'found':False})
            continue
        latest = files[-1]
        # 提取檔名日期
        stem = latest.stem
        date_part = stem.split('_')[-1] if '_' in stem else ''
        if date_part.isdigit() and len(date_part) == 8:
            excel_dates.append(date_part)
        wb = load_workbook(latest, data_only=True)
        ws = wb['Investment Portfolio']
        i25 = ws['I25'].value or 0
        i27 = ws['I27'].value
        i29 = ws['I29'].value
        # 若公式快取為空，從持倉欄位計算
        if i27 is None:
            cash = ws['C25'].value or 0
            stock_val = 0
            for r in range(4, 21):
                qty   = ws.cell(r, 5).value
                price = ws.cell(r, 9).value
                if qty and price:
                    stock_val += qty * price * 1000
            i27 = cash + stock_val
        if i29 is None:
            i29 = i27 - i25
        wb.close()
        # init=None 表示以 I25 為初始投資額（如老哥、Jeff 自營）
        effective_init = i25 if c['init'] is None else c['init']
        data.append({**c, 'aum':i27, 'i25':i25, 'pnl':i29, 'found':True, 'init':effective_init})

    excel_date_str = max(excel_dates) if excel_dates else today_fn

    total_aum = sum(d['aum'] for d in data)
    total_i25 = sum(d['i25'] for d in data if d['found'])
    total_pnl = sum(d['pnl'] for d in data if d['found'])

    wavg_period = (total_pnl / total_i25 * 100) if total_i25 else 0
    wavg_cumul  = ((total_aum - total_i25) / total_i25 * 100) if total_i25 else 0

    for d in data:
        d['share'] = d['aum'] / total_aum * 100 if total_aum else 0
        d['period_ret'] = d['pnl'] / d['i25'] * 100 if d['i25'] else 0
        d['cumul_ret'] = ((d['aum'] / d['i25'] - 1) * 100) if d['i25'] else 0

    n_clients = len(data)

    # ── 文字摘要 ──
    print(f"總 AUM: {total_aum:,.0f}")
    print(f"加權本期報酬率: {wavg_period:+.2f}%")
    print(f"加權累計報酬率: {wavg_cumul:+.2f}%")
    for d in data:
        print(f"  {d['name']}: AUM={d['aum']:,.0f} ({d['share']:.1f}%), "
              f"本期={d['period_ret']:+.2f}%, 累計={d['cumul_ret']:+.2f}%")

    # ══════════════════════ PDF ══════════════════════
    fig = plt.figure(figsize=(13.5, 9.5), facecolor=BG)
    from matplotlib.ticker import FuncFormatter

    total_growth   = total_aum - total_i25
    total_growth_r = total_growth / total_i25 * 100 if total_i25 else 0

    # ── 標題列 ──
    ax_t = fig.add_axes([0, 0.930, 1, 0.070])
    ax_t.set_xlim(0,1); ax_t.set_ylim(0,1); ax_t.axis('off')
    ax_t.set_facecolor(WINE)
    ax_t.text(0.03, 0.50, '全客戶 AUM 彙總報告', color=GOLD, fontsize=22,
              fontweight='bold', va='center', fontproperties=heiti)
    ax_t.text(0.97, 0.50, f'{today_str}  |  客戶數：{n_clients} 位',
              color=WARM, fontsize=11, va='center', ha='right', fontproperties=heiti)

    # ── KPI 卡片 ──
    kpis = [
        ('當前總 AUM',  f'NT$ {total_aum:,.0f}',           GOLD),
        ('初始總 AUM',  f'NT$ {total_i25:,.0f}',           WARM),
        ('總成長金額',  f'NT$ +{total_growth:,.0f}',        color_val(total_growth)),
        ('加權成長率',  f'{total_growth_r:+.2f}%',          color_val(total_growth_r)),
    ]
    card_w = 0.22
    gap = (1 - 4*card_w) / 5
    for i, (label, val, clr) in enumerate(kpis):
        x = gap + i * (card_w + gap)
        ax_k = fig.add_axes([x, 0.820, card_w, 0.095])
        ax_k.set_xlim(0,1); ax_k.set_ylim(0,1)
        ax_k.set_xticks([]); ax_k.set_yticks([])
        ax_k.set_facecolor(CARD_BG)
        for side in ax_k.spines:
            ax_k.spines[side].set_color(BORDER)
            ax_k.spines[side].set_linewidth(2.5)
        ax_k.axhline(y=0.97, xmin=0.02, xmax=0.98, color=GOLD, linewidth=3)
        ax_k.text(0.5, 0.55, val,   color=clr,  fontsize=16, fontweight='bold',
                  ha='center', va='center', fontproperties=heiti)
        ax_k.text(0.5, 0.18, label, color=GRAY, fontsize=10,
                  ha='center', va='center', fontproperties=heiti)

    # ── 圓餅圖（左）— 當前 AUM 佔比 ──
    ax_pie = fig.add_axes([0.04, 0.355, 0.42, 0.415])
    ax_pie.set_facecolor(BG)
    sizes  = [d['aum'] for d in data]
    labels = [f"{d['name']}\n{d['aum']/1e6:.2f}M" for d in data]
    colors = PIE_C[:len(data)]
    wedges, texts = ax_pie.pie(sizes, labels=labels, colors=colors, startangle=90,
                               textprops={'color':WARM, 'fontsize':10},
                               pctdistance=0.75, labeldistance=1.15)
    for t in texts:
        t.set_fontproperties(heiti)
    for i, w in enumerate(wedges):
        ang = (w.theta2 + w.theta1) / 2
        x_p = 0.65 * np.cos(np.radians(ang))
        y_p = 0.65 * np.sin(np.radians(ang))
        ax_pie.text(x_p, y_p, f"{data[i]['share']:.1f}%", color=WARM, fontsize=11,
                    fontweight='bold', ha='center', va='center', fontproperties=heiti)
    ax_pie.set_title('當前 AUM 佔比', color=LGOLD, fontsize=14, fontweight='bold',
                     fontproperties=heiti, pad=10)

    # ── 右：初始 vs 當前 AUM 群組橫條 ──
    ax_bar = fig.add_axes([0.52, 0.370, 0.44, 0.390])
    ax_bar.set_facecolor(CARD_BG)
    for side in ax_bar.spines:
        ax_bar.spines[side].set_color(BORDER)
        ax_bar.spines[side].set_linewidth(2.5)

    sorted_data = sorted(data, key=lambda d: d['cumul_ret'])   # 低→高，圖表上方最高
    names_s    = [d['name']       for d in sorted_data]
    vals_i25_s = [d['i25'] / 1e6  for d in sorted_data]
    vals_i27_s = [d['aum'] / 1e6  for d in sorted_data]

    y = np.arange(len(names_s))
    h = 0.32
    bars1 = ax_bar.barh(y - h/2, vals_i25_s, h, color='#3A5A7A', alpha=0.85, label='初始 AUM')
    bars2 = ax_bar.barh(y + h/2, vals_i27_s, h, color=GOLD,     alpha=0.88, label='當前 AUM')

    ax_bar.set_yticks(y)
    ax_bar.set_yticklabels(names_s, color=WARM, fontsize=11)
    for lbl in ax_bar.get_yticklabels():
        lbl.set_fontproperties(heiti)
    ax_bar.tick_params(axis='x', colors=GRAY, labelsize=9)
    ax_bar.xaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{x:.1f}M'))

    max_val = max(vals_i27_s) if vals_i27_s else 1
    ax_bar.set_xlim(0, max_val * 1.60)

    for i, d in enumerate(sorted_data):
        growth = d['aum'] - d['i25']
        ret    = d['cumul_ret']
        clr    = color_val(ret)
        # 初始 AUM 金額標籤（bar 內置中）
        b1 = bars1[i]
        ax_bar.text(b1.get_width() / 2, b1.get_y() + b1.get_height() / 2,
                    f'{d["i25"]/1e6:.2f}M', color=WARM, fontsize=8,
                    ha='center', va='center', fontproperties=heiti)
        # 當前 AUM 金額標籤（bar 內置中）
        b2 = bars2[i]
        ax_bar.text(b2.get_width() / 2, b2.get_y() + b2.get_height() / 2,
                    f'{d["aum"]/1e6:.2f}M', color='#1A1212', fontsize=8,
                    ha='center', va='center', fontproperties=heiti)
        # 右側成長標籤
        ax_bar.text(b2.get_width() + max_val * 0.03,
                    b2.get_y() + b2.get_height() / 2,
                    f'+{growth/1e6:.2f}M  ({ret:+.1f}%)',
                    color=clr, fontsize=9, va='center', fontproperties=heiti)

    ax_bar.legend(loc='lower right', fontsize=9, facecolor=CARD_BG, edgecolor=BORDER,
                  labelcolor=WARM, prop=heiti)
    ax_bar.set_title('初始 vs 當前 AUM 成長', color=LGOLD, fontsize=14, fontweight='bold',
                     fontproperties=heiti, pad=10)

    # ── 客戶明細卡（底部）— 初始AUM / 當前AUM / 成長率 ──
    card_total_w = 0.92
    card_each_w  = card_total_w / n_clients
    card_gap     = 0.02
    card_actual_w = card_each_w - card_gap
    x_start = (1 - card_total_w) / 2

    for i, d in enumerate(data):
        x = x_start + i * card_each_w
        ax_c = fig.add_axes([x, 0.040, card_actual_w, 0.290])
        ax_c.set_xlim(0,1); ax_c.set_ylim(0,1)
        ax_c.set_xticks([]); ax_c.set_yticks([])
        ax_c.set_facecolor(CARD_BG)
        for side in ax_c.spines:
            ax_c.spines[side].set_color(BORDER)
            ax_c.spines[side].set_linewidth(2.5)

        # 頂部裝飾線：正成長紅、負成長綠（台灣慣例）
        deco_clr = color_val(d['cumul_ret'])
        ax_c.axhline(y=0.97, xmin=0.02, xmax=0.98, color=deco_clr, linewidth=4)

        # 客戶名
        ax_c.text(0.5, 0.82, d['name'], color=WARM, fontsize=16, fontweight='bold',
                  ha='center', va='center', fontproperties=heiti)
        # 類型
        ax_c.text(0.5, 0.70, d['type'], color=GRAY, fontsize=10,
                  ha='center', va='center', fontproperties=heiti)

        # 初始 AUM（I25）
        ax_c.text(0.08, 0.52, '初始 AUM', color=GRAY, fontsize=9.5, ha='left', va='center',
                  fontproperties=heiti)
        ax_c.text(0.92, 0.52, f'{d["i25"]:,.0f}', color=WARM, fontsize=11,
                  ha='right', va='center', fontproperties=heiti)

        # 當前 AUM（I27）
        ax_c.text(0.08, 0.35, '當前 AUM', color=GRAY, fontsize=9.5, ha='left', va='center',
                  fontproperties=heiti)
        ax_c.text(0.92, 0.35, f'{d["aum"]:,.0f}', color=GOLD, fontsize=13,
                  fontweight='bold', ha='right', va='center', fontproperties=heiti)

        # 成長率
        gr_clr = color_val(d['cumul_ret'])
        ax_c.text(0.08, 0.17, '當前收益率', color=GRAY, fontsize=9.5, ha='left', va='center',
                  fontproperties=heiti)
        ax_c.text(0.92, 0.17, f'{d["cumul_ret"]:+.2f}%', color=gr_clr, fontsize=14,
                  fontweight='bold', ha='right', va='center', fontproperties=heiti)

    # ── 頁尾 ──
    ax_f = fig.add_axes([0, 0.000, 1, 0.032])
    ax_f.set_xlim(0,1); ax_f.set_ylim(0,1); ax_f.axis('off')
    ax_f.set_facecolor(BG)
    ax_f.text(0.5, 0.5,
              f'資料截至各客戶 Excel 最新日期（{excel_date_str}）  |  製表日期 {today_str}',
              color=GRAY, fontsize=9, ha='center', va='center', fontproperties=heiti)

    # ── 存檔 ──
    pdf_path = out_dir / f'客戶總覽_{today_fn}.pdf'
    fig.savefig(pdf_path, facecolor=BG, dpi=150, bbox_inches='tight')
    plt.close()
    print(f'\n[PDF] → {pdf_path}')

    return data, total_aum, wavg_period, wavg_cumul

if __name__ == '__main__':
    run()
