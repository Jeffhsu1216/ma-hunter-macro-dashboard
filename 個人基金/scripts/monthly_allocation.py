import openpyxl, glob, math, os
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm

font_path = '/System/Library/Fonts/STHeiti Medium.ttc'
fp = fm.FontProperties(fname=font_path)
fpb = fm.FontProperties(fname=font_path, weight='bold')
fps = fm.FontProperties(fname=font_path, size=8)

# ---------- 1. Read selected stocks ----------
wb_s = openpyxl.load_workbook('/Users/jeffhsu/Desktop/Stock Analysis/Jeff_Stock Analysis.xlsx', data_only=True)
ws_s = wb_s['202604']
stocks = []
for row in range(2, 42):
    if str(ws_s.cell(row=row, column=14).value).strip() == '是':
        code = str(ws_s.cell(row=row, column=2).value).strip()
        name = str(ws_s.cell(row=row, column=3).value).strip()
        price = float(ws_s.cell(row=row, column=8).value)
        stocks.append((code, name, price))
N = len(stocks)
sorder = [s[0] for s in stocks]
sdict = {s[0]: (s[1], s[2]) for s in stocks}

# ---------- 2. Read client data ----------
clients_dir = '/Users/jeffhsu/Desktop/Clients/Excel'
client_cfgs = [
    ('小阿姨', '小阿姨_'),
    ('林峻毅', '林峻毅_'),
    ('靜怡', '靜怡_'),
    ('自營', 'Jeff_'),
]
ODD_MIN = 500

def read_client(cname, prefix):
    import re
    files = sorted(f for f in glob.glob(f'{clients_dir}/{prefix}*.xlsx')
                   if re.search(r'_\d{8}\.xlsx$', f))
    fpath = files[-1]
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb['Investment Portfolio']
    cash = float(ws['C25'].value)
    holdings = {}
    non_selected = []
    for row in range(4, 20):
        code_val = ws.cell(row=row, column=4).value
        if code_val is None:
            continue
        code = str(code_val).strip()
        lots = ws.cell(row=row, column=5).value
        if lots is None:
            continue
        lots = float(lots)
        shares = int(round(lots * 1000))
        cur_p = float(ws.cell(row=row, column=9).value)
        name_val = str(ws.cell(row=row, column=3).value).strip()
        if code in sdict:
            if code in holdings:
                holdings[code]['shares'] += shares
            else:
                holdings[code] = {'shares': shares, 'cur_price': cur_p, 'name': name_val}
        else:
            non_selected.append({'code': code, 'name': name_val, 'shares': shares, 'cur_price': cur_p})
    return cash, holdings, non_selected

def allocate(cash, holdings):
    cur_vals = {}
    for code in sorder:
        if code in holdings:
            cur_vals[code] = holdings[code]['shares'] * holdings[code]['cur_price']
        else:
            cur_vals[code] = 0
    total_inv = sum(cur_vals.values()) + cash
    target = total_inv / N
    result = []
    for code, name, price in stocks:
        cur_val = cur_vals[code]
        cur_sh = holdings[code]['shares'] if code in holdings else 0
        gap = target - cur_val
        buy_lots = 0; buy_odd = 0
        if gap > 0:
            if price >= ODD_MIN:
                total_shares = math.floor(gap / price)
                buy_lots = total_shares // 1000
                buy_odd = total_shares % 1000
            else:
                buy_lots = math.floor(gap / (price * 1000))
                buy_odd = 0
                if cur_sh == 0 and buy_lots == 0:
                    buy_lots = 1
        buy_shares = buy_lots * 1000 + buy_odd
        cost = buy_shares * price
        result.append({
            'code': code, 'name': name, 'price': price,
            'cur_shares': cur_sh, 'cur_val': cur_val,
            'target': target, 'gap': gap,
            'buy_lots': buy_lots, 'buy_odd': buy_odd,
            'buy_shares': buy_shares, 'cost': cost,
        })
    return result, total_inv, target

def should_skip(cash, result):
    total_buy = sum(r['cost'] for r in result)
    if total_buy <= cash:
        return False, ""
    shortfall = total_buy - cash
    if shortfall <= cash * 3:
        return False, ""
    underweight_lot_costs = [r['price']*1000 for r in result if r['gap'] > 0 and r['price'] < ODD_MIN]
    if not underweight_lot_costs or cash >= min(underweight_lot_costs):
        return False, ""
    return True, f"現金僅 {cash:,.0f} 元，需補入 {shortfall:,.0f} 元"

# ---------- Colors ----------
C_TITLE = '#1B2A4A'
C_HEADER_BG = '#1B2A4A'
C_HEADER_FG = 'white'
C_BUY = '#1A7A3A'
C_SKIP_BG = '#FFF0F0'
C_SKIP_BORDER = '#CC3333'
C_SKIP_TEXT = '#CC3333'
C_OVER = '#888888'
C_ROW_ALT = '#F4F6F9'
C_LINE = '#CCCCCC'
C_ORANGE = '#D4760A'
C_DARK = '#222222'
C_SUB = '#666666'

# ---------- Draw one client card on given ax ----------
def draw_card(ax, cname, prefix):
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis('off')

    cash, holdings, non_selected = read_client(cname, prefix)
    result, total_inv, target = allocate(cash, holdings)
    skip, skip_msg = should_skip(cash, result)
    total_buy = sum(r['cost'] for r in result)
    remaining = cash - total_buy
    hold_val = sum(r['cur_val'] for r in result)

    # === Title ===
    ax.text(0.50, 0.96, cname, ha='center', va='top',
            fontproperties=fpb, fontsize=15, color=C_TITLE)

    # === Subtitle: cash / hold / investable ===
    sub = f"現金 {cash:,.0f}   |   持股市值 {hold_val:,.0f}   |   可投資 {total_inv:,.0f}   |   目標 {target:,.0f} / 支"
    ax.text(0.50, 0.90, sub, ha='center', va='top',
            fontproperties=fp, fontsize=8, color=C_SUB)

    if skip:
        # === Skip card ===
        from matplotlib.patches import FancyBboxPatch
        box = FancyBboxPatch((0.10, 0.25), 0.80, 0.55, boxstyle="round,pad=0.02",
                             facecolor=C_SKIP_BG, edgecolor=C_SKIP_BORDER, linewidth=2.5,
                             transform=ax.transAxes)
        ax.add_patch(box)
        ax.text(0.50, 0.60, '本月暫不配置', ha='center', va='center',
                fontproperties=fpb, fontsize=18, color=C_SKIP_TEXT)
        ax.text(0.50, 0.50, skip_msg, ha='center', va='center',
                fontproperties=fp, fontsize=11, color='#993333')
        held = [(r['name'], r['cur_shares'], r['cur_val']) for r in result if r['cur_shares'] > 0]
        if held:
            parts = []
            for n, s, v in held:
                lo = s // 1000
                od = s % 1000
                disp = f"{lo}張" if od == 0 else (f"{lo}張{od}股" if lo > 0 else f"{od}股")
                parts.append(f"{n} {disp}")
            ax.text(0.50, 0.38, '現有持股：' + '、'.join(parts), ha='center', va='center',
                    fontproperties=fp, fontsize=9, color='#888')
        return

    # === Table ===
    headers = ['股票', '現持股', '現值', '目標', '差額', '建議買入', '成本']
    col_x =   [0.02,  0.19,   0.33,  0.46,  0.58,  0.72,    0.88]
    aligns =  ['left','center','right','right','right','center','right']

    y_top = 0.84
    row_h = 0.072

    # Header background
    ax.add_patch(plt.Rectangle((0.01, y_top - 0.005), 0.98, 0.04,
                               transform=ax.transAxes, facecolor=C_HEADER_BG, edgecolor='none'))
    for j, h in enumerate(headers):
        ax.text(col_x[j] + (0.06 if aligns[j] == 'right' else 0),
                y_top + 0.012, h,
                ha=aligns[j] if aligns[j] != 'center' else 'center',
                va='center', fontproperties=fpb, fontsize=8.5, color=C_HEADER_FG,
                transform=ax.transAxes)

    for i, r in enumerate(result):
        y = y_top - (i + 1) * row_h
        # Alternating row bg
        if i % 2 == 0:
            ax.add_patch(plt.Rectangle((0.01, y - 0.02), 0.98, row_h,
                                       transform=ax.transAxes, facecolor=C_ROW_ALT, edgecolor='none'))

        # Current shares display
        if r['cur_shares'] > 0:
            lo = r['cur_shares'] // 1000
            od = r['cur_shares'] % 1000
            cur_disp = f"{lo}張" + (f"+{od}股" if od else "")
        else:
            cur_disp = "—"

        # Buy display
        if r['buy_shares'] > 0:
            parts = []
            if r['buy_lots'] > 0: parts.append(f"{r['buy_lots']}張")
            if r['buy_odd'] > 0: parts.append(f"{r['buy_odd']}股")
            buy_disp = '+'.join(parts)
            buy_color = C_BUY
            cost_color = C_BUY
        elif r['gap'] <= 0:
            buy_disp = '已達標'
            buy_color = C_OVER
            cost_color = C_OVER
        else:
            buy_disp = '—'
            buy_color = C_DARK
            cost_color = C_DARK

        gap_val = r['gap']
        gap_str = f"+{gap_val:,.0f}" if gap_val > 0 else f"{gap_val:,.0f}"
        gap_color = C_BUY if gap_val > 0 else C_OVER

        vals = [
            (f"{r['name']}({r['code']})", C_DARK),
            (cur_disp, C_DARK),
            (f"{r['cur_val']:,.0f}", C_DARK),
            (f"{r['target']:,.0f}", C_SUB),
            (gap_str, gap_color),
            (buy_disp, buy_color),
            (f"{r['cost']:,.0f}" if r['cost'] > 0 else '—', cost_color),
        ]
        for j, (v, c) in enumerate(vals):
            ax.text(col_x[j] + (0.06 if aligns[j] == 'right' else 0),
                    y, v, ha=aligns[j] if aligns[j] != 'center' else 'center',
                    va='center', fontproperties=fp, fontsize=8, color=c,
                    transform=ax.transAxes)

    # === Summary line ===
    y_sum = y_top - (N + 1) * row_h - 0.015
    ax.plot([0.02, 0.98], [y_sum + 0.025, y_sum + 0.025],
            transform=ax.transAxes, color=C_LINE, linewidth=0.8)

    ax.text(0.02, y_sum, f'合計買入：{total_buy:,.0f} 元', transform=ax.transAxes,
            fontproperties=fpb, fontsize=9.5, color=C_BUY)
    rem_color = C_ORANGE if remaining < 100000 else C_DARK
    ax.text(0.55, y_sum, f'剩餘現金：{remaining:,.0f} 元', transform=ax.transAxes,
            fontproperties=fpb, fontsize=9.5, color=rem_color)

    # Non-selected
    if non_selected:
        y_ns = y_sum - 0.045
        ns_parts = [f"{h['name']}({h['code']}) {h['shares']}股 @{h['cur_price']}" for h in non_selected]
        ax.text(0.02, y_ns, '非選股持倉：' + '、'.join(ns_parts),
                transform=ax.transAxes, fontproperties=fp, fontsize=7.5, color='#999')

# ---------- Build PDF ----------
fig = plt.figure(figsize=(20, 24))

# Main title area
fig.text(0.50, 0.975, '月選股下單建議 — 202604（含持倉分析）',
         ha='center', va='top', fontproperties=fpb, fontsize=22, color=C_TITLE)
stock_label = ' | '.join([f'{s[1]}({s[0]}) {s[2]}元' for s in stocks])
fig.text(0.50, 0.958, stock_label,
         ha='center', va='top', fontproperties=fp, fontsize=10, color=C_SUB)
fig.text(0.50, 0.945, '股價 >= 500 可買零股  |  股價 < 500 僅買整張  |  優先買齊 N 支 > Equal Weight',
         ha='center', va='top', fontproperties=fp, fontsize=9, color='#999')

# 2x2 grid with generous spacing
positions = [
    [0.03, 0.50, 0.46, 0.42],   # top-left
    [0.52, 0.50, 0.46, 0.42],   # top-right
    [0.03, 0.03, 0.46, 0.42],   # bottom-left
    [0.52, 0.03, 0.46, 0.42],   # bottom-right
]

for idx, (cname, prefix) in enumerate(client_cfgs):
    ax = fig.add_axes(positions[idx])
    # Card border
    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_color('#DDDDDD')
        spine.set_linewidth(1)
    draw_card(ax, cname, prefix)

out_path = '/Users/jeffhsu/Desktop/Clients/AUM_Overview/月選股下單建議_202604_含持倉_20260408.pdf'

from matplotlib.backends.backend_pdf import PdfPages
with PdfPages(out_path) as pdf:
    pdf.savefig(fig, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig)

    # ========== Page 2: Order Summary ==========
    fig2 = plt.figure(figsize=(20, 24))
    fig2.text(0.50, 0.97, '下單總結 — 202604', ha='center', va='top',
              fontproperties=fpb, fontsize=24, color=C_TITLE)
    fig2.text(0.50, 0.955, '各帳戶本月應下單明細（僅列有下單之帳戶）',
              ha='center', va='top', fontproperties=fp, fontsize=12, color=C_SUB)

    # Collect all order data
    all_orders = []
    total_all = 0
    for cname, prefix in client_cfgs:
        cash, holdings, non_selected = read_client(cname, prefix)
        result, total_inv, target = allocate(cash, holdings)
        skip, skip_msg = should_skip(cash, result)
        buys = [r for r in result if r['buy_shares'] > 0]
        total_buy = sum(r['cost'] for r in result)
        remaining = cash - total_buy
        all_orders.append({
            'name': cname, 'cash': cash, 'skip': skip, 'skip_msg': skip_msg,
            'buys': buys, 'total_buy': total_buy, 'remaining': remaining,
        })
        if not skip:
            total_all += total_buy

    # --- Draw summary table per client ---
    ax2 = fig2.add_axes([0.08, 0.08, 0.84, 0.84])
    ax2.set_xlim(0, 1)
    ax2.set_ylim(0, 1)
    ax2.axis('off')

    y = 0.95
    grand_total = 0

    for oi, od in enumerate(all_orders):
        if od['skip']:
            continue

        # Client header bar
        ax2.add_patch(plt.Rectangle((0.0, y - 0.005), 1.0, 0.045,
                      transform=ax2.transAxes, facecolor=C_HEADER_BG, edgecolor='none'))
        status = f"下單 {len(od['buys'])} 筆  |  合計 {od['total_buy']:,.0f} 元  |  剩餘現金 {od['remaining']:,.0f} 元"
        ax2.text(0.02, y + 0.015, od['name'], ha='left', va='center',
                 fontproperties=fpb, fontsize=14, color=C_HEADER_FG, transform=ax2.transAxes)
        ax2.text(0.98, y + 0.015, status, ha='right', va='center',
                 fontproperties=fp, fontsize=11, color=C_HEADER_FG, transform=ax2.transAxes)
        y -= 0.055

        if True:
            # Column headers for orders
            oh = ['股票代號', '股票名稱', '股價', '買入方式', '股數', '預估成本']
            ox = [0.04,      0.16,      0.32,   0.46,      0.64,   0.82]
            oa = ['left',    'left',    'right', 'center',  'center','right']
            for j, h in enumerate(oh):
                xoff = 0.08 if oa[j] == 'right' else 0
                ax2.text(ox[j] + xoff, y + 0.008, h, ha=oa[j], va='center',
                         fontproperties=fpb, fontsize=10, color=C_SUB, transform=ax2.transAxes)
            y -= 0.035

            for bi, b in enumerate(od['buys']):
                # Alternating row
                if bi % 2 == 0:
                    ax2.add_patch(plt.Rectangle((0.02, y - 0.01), 0.96, 0.035,
                                  transform=ax2.transAxes, facecolor=C_ROW_ALT, edgecolor='none'))

                # Format buy method
                if b['buy_lots'] > 0 and b['buy_odd'] > 0:
                    method = f"{b['buy_lots']}張 + {b['buy_odd']}股零股"
                elif b['buy_lots'] > 0:
                    method = f"整張"
                else:
                    method = f"零股"

                if b['buy_lots'] > 0 and b['buy_odd'] > 0:
                    shares_disp = f"{b['buy_lots']}張+{b['buy_odd']}股"
                elif b['buy_lots'] > 0:
                    shares_disp = f"{b['buy_lots']}張"
                else:
                    shares_disp = f"{b['buy_odd']}股"

                vals = [
                    (b['code'], C_DARK),
                    (b['name'], C_DARK),
                    (f"{b['price']:,.1f}", C_SUB),
                    (method, C_BUY),
                    (shares_disp, C_BUY),
                    (f"{b['cost']:,.0f}", C_BUY),
                ]
                for j, (v, c) in enumerate(vals):
                    xoff = 0.08 if oa[j] == 'right' else 0
                    ax2.text(ox[j] + xoff, y + 0.008, v, ha=oa[j], va='center',
                             fontproperties=fpb if j >= 3 else fp, fontsize=11, color=c,
                             transform=ax2.transAxes)
                y -= 0.035

            grand_total += od['total_buy']

        # Separator
        y -= 0.015
        ax2.plot([0.02, 0.98], [y + 0.01, y + 0.01],
                 transform=ax2.transAxes, color=C_LINE, linewidth=0.5)
        y -= 0.015

    # ===== Grand total box =====
    y -= 0.03
    ax2.add_patch(plt.Rectangle((0.15, y - 0.01), 0.70, 0.05,
                  transform=ax2.transAxes, facecolor='#EBF5EC', edgecolor=C_BUY, linewidth=1.5))
    ax2.text(0.50, y + 0.015, f'全帳戶合計下單金額：{grand_total:,.0f} 元',
             ha='center', va='center', fontproperties=fpb, fontsize=16, color=C_BUY,
             transform=ax2.transAxes)

    # ===== Skip note =====
    skipped = [od for od in all_orders if od['skip']]
    if skipped:
        y -= 0.06
        skip_names = '、'.join([od['name'] for od in skipped])
        ax2.text(0.50, y, f'* {skip_names} 本月暫不配置（資金不足）',
                 ha='center', va='center', fontproperties=fp, fontsize=11, color=C_OVER,
                 transform=ax2.transAxes)

    pdf.savefig(fig2, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close(fig2)

print(f"PDF saved: {out_path}")
