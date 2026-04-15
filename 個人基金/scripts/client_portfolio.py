import openpyxl, time, os, sys, glob
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.ticker as mticker
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.patches import FancyBboxPatch, Rectangle
from datetime import datetime

t0 = time.time()

# ── 客戶設定表 ──
CLIENT_MAP = {
    '林峻毅': {'type': 'Fee-Based', 'folder': 'Fee-Based_林峻毅',  'prefix': '林峻毅_', 'init': 2_000_000},
    '小阿姨': {'type': 'Non-Fee',   'folder': 'Non-Fee_小阿姨',    'prefix': '小阿姨_', 'init': 1_318_502},
    '靜怡':   {'type': 'Fee-Based', 'folder': 'Fee-Based_靜怡',    'prefix': '靜怡_',   'init': 600_000},
    'Jeff':   {'type': '自營',      'folder': '自營_Jeff',          'prefix': 'Jeff_',   'init': 4_233_490},
    'Gary':   {'type': 'Fee-Based', 'folder': 'Fee-Based_Gary',    'prefix': 'Gary_',   'init': 4_000_000},
    'Josh':   {'type': 'Non-Fee',   'folder': 'Non-Fee_老哥',       'prefix': 'Josh_',   'init': None},   # 老哥：init=I25
}

# ── 結帳週期（依當月自動判斷）──
def get_period_start():
    now = datetime.today()
    m = now.month
    if 4 <= m <= 9:
        return datetime(now.year, 4, 1)
    elif m <= 3:
        return datetime(now.year - 1, 10, 1)
    else:
        return datetime(now.year, 10, 1)

PERIOD_START = get_period_start()
TODAY = datetime.today().strftime("%Y/%m/%d")
TODAY_FILE = datetime.today().strftime("%Y%m%d")

# ── 配色 ──
BG     = '#141010'
GOLD   = '#C8A951'
LGOLD  = '#D4BC7A'
WHITE  = '#F5F0E8'
GRAY   = '#9E8E7E'
GREEN  = '#4CAF50'
RED    = '#E74C3C'
BLUE   = '#6AAFFF'
DARKBL = '#1A2A4A'
CARD   = '#1A1212'
CARDB  = '#3A2828'
DIVID  = '#5A3A2A'

plt.rcParams['font.family'] = ['PingFang HK','Heiti TC','Arial Unicode MS','DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False


def find_latest_excel(prefix):
    import re
    pattern = os.path.expanduser(f'~/Desktop/Clients/Excel/{prefix}*.xlsx')
    files = [f for f in glob.glob(pattern)
             if re.search(r'_\d{8}\.xlsx$', f)]   # 只取 YYYYMMDD 格式
    if not files:
        return None
    return max(files)


def run_client(name):
    cfg = CLIENT_MAP[name]
    excel_path = find_latest_excel(cfg['prefix'])
    if not excel_path:
        print(f"[ERROR] 找不到 {name} 的 Excel 檔案")
        return

    # ── 讀 Excel ──
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Investment Portfolio']
    i25 = ws['I25'].value or 0
    i27 = ws['I27'].value
    i29 = ws['I29'].value

    # Fallback：I27/I29 為 None 時從持倉欄位計算市值
    if i27 is None or i27 == 0:
        cash = ws['C25'].value or 0
        stock_val = 0
        for r in range(4, 25):
            qty   = ws.cell(r, 5).value   # E欄：張數
            price = ws.cell(r, 9).value   # I欄：當前價格
            if qty and price:
                stock_val += qty * price * 1000
        i27 = cash + stock_val
    if i29 is None:
        i29 = i27 - i25

    INIT = i25   # 初始AUM 統一用 I25

    # ── 已實現交易 ──
    trades = []
    for r in range(4, 200):
        n = ws.cell(r, 14).value
        p = ws.cell(r, 16).value
        if n and p:
            buy_dt = ws.cell(r, 13).value
            if isinstance(buy_dt, str):
                buy_dt = datetime.strptime(buy_dt[:10], '%Y-%m-%d')
            elif hasattr(buy_dt, 'date'):
                buy_dt = datetime(buy_dt.year, buy_dt.month, buy_dt.day)
            if buy_dt is None:
                continue
            trades.append({
                'name': n,
                'buy':  buy_dt,
                'days': ws.cell(r, 17).value or 0,
                'pnl':  ws.cell(r, 24).value or 0,
                'ret':  (ws.cell(r, 25).value or 0) * 100,
            })

    cur = [t for t in trades if t['buy'] >= PERIOD_START]
    stocks = {}
    for t in cur:
        stocks.setdefault(t['name'], 0)
        stocks[t['name']] += t['pnl']
    winners    = sum(1 for v in stocks.values() if v > 0)
    win_rate   = winners / len(stocks) * 100 if stocks else 0
    avg_days   = sum(t['days'] for t in cur) / len(cur) if cur else 0
    avg_ret    = sum(t['ret'] for t in cur) / len(cur) if cur else 0
    period_ret = (i29 / i25 * 100) if i25 else 0

    # 本期月數：從 period_start 到今日
    months_elapsed = max(1, (datetime.today().year - PERIOD_START.year) * 12
                         + datetime.today().month - PERIOD_START.month)
    monthly_ret = period_ret / months_elapsed
    cum_ret     = ((i27 / INIT - 1) * 100) if INIT else 0
    best  = max(cur, key=lambda t: t['pnl']) if cur else None
    worst = min(cur, key=lambda t: t['pnl']) if cur else None
    top3  = sorted([(s, v) for s, v in stocks.items()], key=lambda x: -x[1])[:3]
    names_cur = [t['name'] for t in cur]
    most_traded = max(set(names_cur), key=names_cur.count) if names_cur else '-'

    period_label = (f"{PERIOD_START.year}/{PERIOD_START.month:02d} ～ "
                    f"{datetime.today().year}/{datetime.today().month:02d}")

    # ── 畫布 ──
    fig = plt.figure(figsize=(13.33, 9.5), facecolor=BG)
    gs  = gridspec.GridSpec(4, 1, figure=fig,
          height_ratios=[0.068, 0.068, 0.50, 0.364],
          left=0.03, right=0.97, top=0.97, bottom=0.035,
          hspace=0.035)

    def no_ax(ax, bg=BG):
        ax.set_facecolor(bg); ax.set_xticks([]); ax.set_yticks([])
        for sp in ax.spines.values(): sp.set_visible(False)

    # ══ 1. 標題 ══
    ax_t = fig.add_subplot(gs[0]); no_ax(ax_t)
    ax_t.plot([0,1],[0,0], color=GOLD, lw=1.2, transform=ax_t.transAxes, clip_on=False)
    ax_t.text(0.01, 0.52, f'{name}  資產現況報告',
        color=GOLD, fontsize=17, fontweight='bold', va='center', transform=ax_t.transAxes)
    ax_t.text(0.99, 0.52, f'{cfg["type"]}  |  {TODAY}',
        color=GRAY, fontsize=9.5, va='center', ha='right', transform=ax_t.transAxes)

    # ══ 2. KPI 欄 ══
    ax_k = fig.add_subplot(gs[1]); no_ax(ax_k, bg='#1C1212')
    ax_k.plot([0,1],[0,0], color=GOLD, lw=0.6, alpha=0.5, transform=ax_k.transAxes, clip_on=False)
    ax_k.plot([0,1],[1,1], color=GOLD, lw=0.6, alpha=0.5, transform=ax_k.transAxes, clip_on=False)
    kpis = [
        ('期初投資額', f'NT$ {i25:,.0f}',    WHITE),
        ('當前餘額',   f'NT$ {i27:,.0f}',    BLUE),
        ('本期總損益', f'NT$ {i29:+,.0f}',   GREEN if i29 >= 0 else RED),
        ('本期總報酬', f'{period_ret:+.2f}%', GREEN if period_ret >= 0 else RED),
    ]
    for i, (lbl, val, col) in enumerate(kpis):
        x0 = i * 0.25
        if i > 0:
            ax_k.plot([x0,x0],[0,1], color=GOLD, lw=0.8, alpha=0.4,
                transform=ax_k.transAxes, clip_on=False)
        ax_k.text(x0+0.125, 0.78, lbl,
            color=GRAY, fontsize=8, ha='center', va='center', transform=ax_k.transAxes)
        ax_k.text(x0+0.125, 0.30, val,
            color=col, fontsize=12, fontweight='bold', ha='center', va='center',
            transform=ax_k.transAxes)

    # ══ 3. 主體 (左 + 右) ══
    gs_m = gridspec.GridSpecFromSubplotSpec(1, 2, subplot_spec=gs[2],
        wspace=0.03, width_ratios=[1,1])

    # ── 左：交易績效總覽 ──
    axL = fig.add_subplot(gs_m[0]); no_ax(axL, bg=CARD)
    axL.add_patch(Rectangle((0,0),1,1, facecolor=CARD, edgecolor=DIVID, lw=0.6,
        transform=axL.transAxes))
    axL.text(0.50, 0.965, '交易績效總覽',
        color=GOLD, fontsize=12, fontweight='bold', ha='center', va='top',
        transform=axL.transAxes)
    axL.text(0.50, 0.895, f'本期: {period_label}',
        color=GRAY, fontsize=8, ha='center', va='top', transform=axL.transAxes)

    rows = [
        ('總交易次數', f'{len(cur)} 筆', WHITE),
        ('勝率（依個股）', f'{win_rate:.0f}%  ({winners}/{len(stocks)} 檔)',
            GREEN if win_rate >= 50 else RED),
        ('平均持有天數', f'{avg_days:.0f} 天', WHITE),
        ('月平均報酬率', f'{monthly_ret:+.2f}%', GREEN if monthly_ret >= 0 else RED),
    ]
    ytops = [0.84, 0.70, 0.56, 0.42]
    for (lbl, val, col), yt in zip(rows, ytops):
        axL.add_patch(FancyBboxPatch((0.04, yt-0.085), 0.92, 0.115,
            boxstyle='round,pad=0.005', facecolor='#221616', edgecolor=CARDB, lw=0.4,
            transform=axL.transAxes))
        axL.text(0.09, yt-0.025, lbl, color=GRAY, fontsize=8.5, va='center',
            transform=axL.transAxes)
        axL.text(0.95, yt-0.025, val, color=col, fontsize=9, fontweight='bold',
            ha='right', va='center', transform=axL.transAxes)

    if best:
        for bx, (t, bg_c, col2, sym) in enumerate([
            (best,  '#1B3A1B', GREEN, '▲'),
            (worst, '#3A1B1B', RED,   '▼')]):
            xp = 0.04 + bx * 0.49
            axL.add_patch(FancyBboxPatch((xp, 0.03), 0.45, 0.24,
                boxstyle='round,pad=0.01', facecolor=bg_c, edgecolor=col2, lw=0.8,
                transform=axL.transAxes))
            axL.text(xp+0.225, 0.215, f'{sym} {t["name"]}', color=col2, fontsize=9,
                fontweight='bold', ha='center', va='center', transform=axL.transAxes)
            axL.text(xp+0.225, 0.115,
                f'{t["pnl"]:+,.0f}  ({t["ret"]:+.1f}%)',
                color=col2, fontsize=8, ha='center', va='center', transform=axL.transAxes)

    # ── 右：三區塊 ──
    axR = fig.add_subplot(gs_m[1]); no_ax(axR, bg=CARD)
    axR.add_patch(Rectangle((0,0),1,1, facecolor=CARD, edgecolor=DIVID, lw=0.6,
        transform=axR.transAxes))

    axR.text(0.04, 0.980, '投資報酬結構',
        color=GOLD, fontsize=10, fontweight='bold', va='top', transform=axR.transAxes)

    ax_bar = axR.inset_axes([0.04, 0.658, 0.92, 0.262])
    ax_bar.set_facecolor('#0F0C0C')
    for sp in ax_bar.spines.values(): sp.set_visible(False)
    ax_bar.set_xticks([]); ax_bar.set_yticks([])
    ax_bar.bar([0], [i25], color='#7A5C1E', width=0.55)
    ax_bar.bar([0], [i29], color=GREEN if i29 >= 0 else RED,
        bottom=[i25], width=0.55, alpha=0.9)
    top_val = i25 + (i29 if i29 >= 0 else 0)
    ax_bar.set_ylim(0, top_val * 1.055)
    ax_bar.set_xlim(-0.8, 0.8)
    pnl_col = GREEN if i29 >= 0 else RED
    ax_bar.text(0.98, 0.96, f'本期總報酬: {period_ret:+.2f}%',
        color=pnl_col, fontsize=8, ha='right', va='top',
        transform=ax_bar.transAxes)
    step = max(500_000, round(top_val / 4 / 500_000) * 500_000)
    tick_max = int(top_val * 1.055 / step) * step
    for mval in range(0, tick_max + step, step):
        if mval <= top_val * 1.055:
            va = 'bottom' if mval == 0 else 'center'
            ax_bar.text(-0.78, mval, f'{mval/1e6:.1f}M',
                color=GRAY, fontsize=7, ha='left', va=va)

    # ② 獲利排行 Top 3
    axR.text(0.04, 0.630, f'獲利排行 Top 3（本期 {period_label}）',
        color=LGOLD, fontsize=8.5, va='top', transform=axR.transAxes)
    medal_colors = [GOLD, '#C0C0C0', '#CD7F32']
    max_pnl = top3[0][1] if top3 else 1
    for i, (sn, sv) in enumerate(top3):
        y = 0.500 - i * 0.100
        bar_w = max((sv / max_pnl) * 0.68, 0.02) if max_pnl > 0 else 0.02
        axR.add_patch(Rectangle((0.04, y+0.010), 0.025, 0.055,
            facecolor=medal_colors[i], transform=axR.transAxes))
        axR.add_patch(Rectangle((0.10, y+0.010), bar_w, 0.050,
            facecolor=GREEN, alpha=0.85, transform=axR.transAxes))
        axR.text(0.115, y+0.035, sn, color=WHITE, fontsize=9, va='center',
            transform=axR.transAxes)
        axR.text(0.96, y+0.035, f'+{sv:,.0f}', color=GREEN, fontsize=9,
            fontweight='bold', ha='right', va='center', transform=axR.transAxes)

    # ③ 交易概況（2×2）
    axR.text(0.04, 0.255, '交易概況',
        color=LGOLD, fontsize=8.5, va='top', transform=axR.transAxes)
    mini = [
        ('交易筆數', f'{len(cur)} 筆'),
        ('操作個檔', f'{len(stocks)} 檔'),
        ('平均報酬',  f'{avg_ret:+.1f}%'),
        ('最常交易', most_traded),
    ]
    for i, (lbl, val) in enumerate(mini):
        xi = i % 2; yi = i // 2
        xp = 0.04 + xi * 0.485
        yp = 0.130 - yi * 0.118
        axR.add_patch(FancyBboxPatch((xp, yp), 0.46, 0.105,
            boxstyle='round,pad=0.005', facecolor='#221616', edgecolor=CARDB, lw=0.4,
            transform=axR.transAxes))
        axR.text(xp+0.23, yp+0.077, lbl,
            color=GRAY, fontsize=7.5, ha='center', va='center', transform=axR.transAxes)
        axR.text(xp+0.23, yp+0.028, val,
            color=WHITE, fontsize=9, fontweight='bold', ha='center', va='center',
            transform=axR.transAxes)

    # ══ 4. 底部 3 大卡 ══
    gs_b = gridspec.GridSpecFromSubplotSpec(1, 3, subplot_spec=gs[3], wspace=0.03)
    bot = [
        ('初始投資額', f'NT$ {INIT:,.0f}', None,
            '#2A1E08', GOLD),
        ('當前餘額',   f'NT$ {i27:,.0f}', None,
            DARKBL, BLUE),
        ('累計報酬率', f'{cum_ret:+.2f}%',
            f'({INIT:,.0f} → {i27:,.0f})',
            '#1B3A1B' if cum_ret >= 0 else '#3A1B1B',
            GREEN if cum_ret >= 0 else RED),
    ]
    for i, (lbl, val, note, bg_c, col) in enumerate(bot):
        axB = fig.add_subplot(gs_b[i])
        no_ax(axB, bg=bg_c)
        axB.add_patch(Rectangle((0,0),1,1, facecolor=bg_c, edgecolor=GOLD, lw=1.2,
            transform=axB.transAxes))
        axB.plot([0,1],[0.90,0.90], color=GOLD, lw=3.5, transform=axB.transAxes, clip_on=False)
        axB.text(0.50, 0.74, lbl,
            color=GRAY, fontsize=10, ha='center', va='center', transform=axB.transAxes)
        axB.text(0.50, 0.44, val,
            color=col, fontsize=18, fontweight='bold', ha='center', va='center',
            transform=axB.transAxes)
        if note:
            axB.text(0.50, 0.17, note,
                color=GRAY, fontsize=8, ha='center', va='center', transform=axB.transAxes)

    # ── 頁尾 ──
    fig.text(0.50, 0.005, f'資料截至 {TODAY}  |  製表日期 {TODAY}  |  圖靈投資管理',
        color=GRAY, fontsize=7.5, ha='center', va='bottom')

    # ══════════════ 第二頁：持倉明細 ══════════════
    # 讀取持倉（C-K 欄，row 4 起，遇到 C 為空停止）
    holdings = []
    for r in range(4, 30):
        stock_name = ws.cell(r, 3).value   # C: 股票名稱
        if not stock_name:
            break
        code     = ws.cell(r, 4).value or ''   # D: 代號
        qty      = ws.cell(r, 5).value or 0    # E: 張數
        buy_px   = ws.cell(r, 6).value or 0    # F: 買入價
        cost     = ws.cell(r, 7).value or 0    # G: 買入成本
        cur_px   = ws.cell(r, 9).value or 0    # I: 當前價
        pnl      = ws.cell(r, 10).value or 0   # J: 損益
        ret      = (ws.cell(r, 11).value or 0) * 100  # K: 報酬率%
        mkt_val  = qty * cur_px * 1000
        holdings.append(dict(
            name=stock_name, code=str(code), qty=qty,
            buy_px=buy_px, cur_px=cur_px, cost=cost,
            mkt_val=mkt_val, pnl=pnl, ret=ret
        ))
    cash_val = ws['C25'].value or 0

    fig2 = plt.figure(figsize=(13.33, 9.5), facecolor=BG)

    def no_ax2(ax, bg=BG):
        ax.set_facecolor(bg); ax.set_xticks([]); ax.set_yticks([])
        for sp in ax.spines.values(): sp.set_visible(False)

    # ── P2 標題列 ──
    ax2_t = fig2.add_axes([0, 0.935, 1, 0.062])
    no_ax2(ax2_t)
    ax2_t.plot([0,1],[0,0], color=GOLD, lw=1.2, transform=ax2_t.transAxes, clip_on=False)
    ax2_t.text(0.01, 0.50, f'{name}  持倉明細',
        color=GOLD, fontsize=17, fontweight='bold', va='center', transform=ax2_t.transAxes)
    ax2_t.text(0.99, 0.50, f'{cfg["type"]}  |  {TODAY}',
        color=GRAY, fontsize=9.5, va='center', ha='right', transform=ax2_t.transAxes)

    # ══ P2 主體：GridSpec 3 區 ══
    # left=表格 (0.02~0.59)  right_top=橫條圖 (0.62~0.985)  right_bot=圓餅 (0.62~0.985)
    # 使用 GridSpec 管理右側兩塊，避免高度計算錯誤

    # ── P2 左：持倉表格 ──
    ax2_tbl = fig2.add_axes([0.02, 0.055, 0.57, 0.865])
    no_ax2(ax2_tbl, bg=CARD)
    ax2_tbl.add_patch(Rectangle((0,0),1,1, facecolor=CARD, edgecolor=DIVID, lw=0.8,
        transform=ax2_tbl.transAxes))

    # 欄位定義：(label, x_pos, align)
    COL_DEFS = [
        ('股票名稱', 0.01,  'left'),
        ('代號',     0.245, 'center'),
        ('張數',     0.330, 'right'),
        ('買入價',   0.415, 'right'),
        ('當前價',   0.500, 'right'),
        ('買入成本', 0.612, 'right'),
        ('當前市值', 0.730, 'right'),
        ('損益',     0.842, 'right'),
        ('報酬率',   0.960, 'right'),
    ]
    COL_X = [c[1] for c in COL_DEFS]

    n_rows    = len(holdings)
    n_total   = n_rows + 2          # header + data + cash + total
    row_h     = min(0.075, 0.870 / max(n_total + 1, 5))
    hdr_y     = 0.960
    data_top  = hdr_y - row_h       # 第一筆資料的 y

    # 標題列背景
    ax2_tbl.add_patch(Rectangle((0.005, hdr_y - row_h * 0.88), 0.990, row_h * 0.88,
        facecolor='#2A1A10', transform=ax2_tbl.transAxes))
    sep_y = hdr_y - row_h * 0.96
    ax2_tbl.plot([0.005, 0.995], [sep_y, sep_y], color=GOLD, lw=0.8,
        alpha=0.7, transform=ax2_tbl.transAxes)

    for (lbl, xc, align) in COL_DEFS:
        ha = align
        ax2_tbl.text(xc, hdr_y - row_h * 0.42, lbl,
            color=GOLD, fontsize=8.0, fontweight='bold', ha=ha, va='center',
            transform=ax2_tbl.transAxes)

    total_cost = 0; total_mkt = 0; total_pnl_h = 0

    for idx, h in enumerate(holdings):
        ry = data_top - idx * row_h
        row_bg = '#1E1212' if idx % 2 == 0 else '#171010'
        ax2_tbl.add_patch(Rectangle((0.005, ry - row_h * 0.86), 0.990, row_h * 0.86,
            facecolor=row_bg, transform=ax2_tbl.transAxes))

        ret_col = GREEN if h['ret'] >= 0 else RED
        pnl_col = GREEN if h['pnl'] >= 0 else RED
        qty_str = f"{h['qty']:.1f}".rstrip('0').rstrip('.')

        vals = [
            (h['name'],              COL_X[0], 'left',   WHITE,   9.0),
            (h['code'],              COL_X[1], 'center', GRAY,    8.0),
            (qty_str,                COL_X[2], 'right',  WHITE,   8.0),
            (f"{h['buy_px']:,.1f}",  COL_X[3], 'right',  GRAY,    8.0),
            (f"{h['cur_px']:,.1f}",  COL_X[4], 'right',  WHITE,   8.0),
            (f"{h['cost']:,.0f}",    COL_X[5], 'right',  GRAY,    8.0),
            (f"{h['mkt_val']:,.0f}", COL_X[6], 'right',  WHITE,   8.0),
            (f"{h['pnl']:+,.0f}",   COL_X[7], 'right',  pnl_col, 8.5),
            (f"{h['ret']:+.2f}%",   COL_X[8], 'right',  ret_col, 8.5),
        ]
        for (txt, xc, ha, col, fs) in vals:
            ax2_tbl.text(xc, ry - row_h * 0.40, txt,
                color=col, fontsize=fs, ha=ha, va='center',
                transform=ax2_tbl.transAxes)

        total_cost  += h['cost']
        total_mkt   += h['mkt_val']
        total_pnl_h += h['pnl']

    # 現金列
    cash_y = data_top - n_rows * row_h
    ax2_tbl.add_patch(Rectangle((0.005, cash_y - row_h * 0.86), 0.990, row_h * 0.86,
        facecolor='#1A1C1A', transform=ax2_tbl.transAxes))
    ax2_tbl.text(COL_X[0], cash_y - row_h * 0.40, '現金',
        color=LGOLD, fontsize=9.0, ha='left', va='center', transform=ax2_tbl.transAxes)
    ax2_tbl.text(COL_X[6], cash_y - row_h * 0.40, f'{cash_val:,.0f}',
        color=LGOLD, fontsize=8.0, ha='right', va='center', transform=ax2_tbl.transAxes)

    # 合計列
    tot_y = cash_y - row_h
    ax2_tbl.plot([0.005, 0.995], [tot_y + row_h * 0.12, tot_y + row_h * 0.12],
        color=GOLD, lw=0.6, alpha=0.5, transform=ax2_tbl.transAxes)
    ax2_tbl.add_patch(Rectangle((0.005, tot_y - row_h * 0.88), 0.990, row_h * 0.88,
        facecolor='#2A1A10', transform=ax2_tbl.transAxes))
    tot_ret_col = GREEN if total_pnl_h >= 0 else RED
    tot_vals = [
        ('合計',                    COL_X[0], 'left',  GOLD),
        (f'{total_cost:,.0f}',     COL_X[5], 'right', GRAY),
        (f'{total_mkt:,.0f}',      COL_X[6], 'right', WHITE),
        (f'{total_pnl_h:+,.0f}',  COL_X[7], 'right', tot_ret_col),
        (f'{(total_pnl_h/total_cost*100):+.2f}%' if total_cost else '—',
                                    COL_X[8], 'right', tot_ret_col),
    ]
    for (txt, xc, ha, col) in tot_vals:
        ax2_tbl.text(xc, tot_y - row_h * 0.42, txt,
            color=col, fontsize=8.5, fontweight='bold', ha=ha, va='center',
            transform=ax2_tbl.transAxes)

    # ── P2 右上：個股報酬率橫條圖 ──
    # 固定佔右側上半（y: 0.505 ~ 0.920），留 title 空間
    ax2_bar = fig2.add_axes([0.615, 0.505, 0.370, 0.415])
    ax2_bar.set_facecolor('#0F0C0C')
    for sp in ax2_bar.spines.values():
        sp.set_color(DIVID); sp.set_linewidth(0.5)
    ax2_bar.spines['top'].set_visible(False)
    ax2_bar.spines['right'].set_visible(False)

    # 在 axes 外畫標題（避免 title pad 與 bar 重疊）
    fig2.text(0.800, 0.926, '個股報酬率',
        color=GOLD, fontsize=10.5, fontweight='bold', ha='center', va='bottom')

    sorted_h = sorted(holdings, key=lambda h: h['ret'])
    names_s  = [h['name'] for h in sorted_h]
    rets_s   = [h['ret']  for h in sorted_h]
    colors_s = [GREEN if r >= 0 else RED for r in rets_s]

    n_s   = len(sorted_h)
    bar_h = min(0.52, 3.8 / max(n_s, 1))   # 動態 bar 高度
    y_pos = list(range(n_s))
    bars  = ax2_bar.barh(y_pos, rets_s, color=colors_s, alpha=0.85, height=bar_h)
    ax2_bar.axvline(0, color=GRAY, lw=0.8, alpha=0.5)

    # ★ 不使用 set_yticklabels（會溢出圖外），改用 axes座標渲染在圖內
    ax2_bar.set_yticks([])
    ax2_bar.tick_params(axis='x', colors=GRAY, labelsize=7.5)
    ax2_bar.xaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f'{v:+.0f}%'))

    # xlim：留足夠空間給數值標注，不越界
    max_abs = max(abs(r) for r in rets_s) if rets_s else 1
    pad     = max_abs * 0.38
    ax2_bar.set_xlim(-max_abs - pad, max_abs + pad)
    ax2_bar.set_ylim(-0.6, n_s - 0.4)

    # 股票名稱：用 axes 座標放在左邊界內，永不溢出
    name_fs = max(6.5, min(8.5, 70 / max(n_s, 1)))
    for i, sname in enumerate(names_s):
        y_ax = (i + 0.5) / n_s          # axes 座標 y（0=底, 1=頂）
        ax2_bar.text(0.015, y_ax, sname,
            color=WHITE, fontsize=name_fs, va='center', ha='left',
            transform=ax2_bar.transAxes, clip_on=True)

    # 數值標注：緊貼 bar 端
    for bar, ret in zip(bars, rets_s):
        offset = max_abs * 0.04
        ha_ann = 'left' if ret >= 0 else 'right'
        xpos   = ret + (offset if ret >= 0 else -offset)
        ax2_bar.text(xpos, bar.get_y() + bar.get_height()/2,
            f'{ret:+.2f}%',
            color=GREEN if ret >= 0 else RED,
            fontsize=7.0, va='center', ha=ha_ann, fontweight='bold', clip_on=True)

    # ── P2 右下：持倉成本圓餅（含現金）──
    # 圓餅佔上 75%，下 25% 留給圖例（在 axes 座標內渲染）
    ax2_pie = fig2.add_axes([0.615, 0.060, 0.370, 0.420])
    ax2_pie.set_facecolor('#0F0C0C')
    for sp in ax2_pie.spines.values(): sp.set_visible(False)

    fig2.text(0.800, 0.487, '持倉成本佔比（含現金）',
        color=GOLD, fontsize=10.5, fontweight='bold', ha='center', va='bottom')

    PIE_COLORS = ['#C8421A','#C8A951','#6A5ACD','#2E8B8B','#5B8B3A',
                  '#4A7FAA','#8B6A1A','#7A2A6A','#3A7A5A','#AA6A2A',
                  '#2A5A8A','#8A2A2A']
    pie_labels = [h['name'] for h in holdings] + ['現金']
    pie_sizes  = [h['cost'] for h in holdings] + [cash_val]
    pie_sizes  = [max(v, 0) for v in pie_sizes]
    pie_colors = (PIE_COLORS * 2)[:len(pie_sizes)]
    total_pie  = sum(pie_sizes)

    if total_pie > 0:
        # 圓餅只佔 axes 上半（center=(0.5, 0.62), radius 不超出）
        wedges, _, autotexts = ax2_pie.pie(
            pie_sizes,
            labels=None,
            colors=pie_colors,
            autopct='%1.1f%%',
            startangle=90,
            pctdistance=0.75,
            center=(0.5, 0.62),      # 在 axes data 單位中偏上
            radius=0.38,
            wedgeprops={'linewidth': 0.4, 'edgecolor': BG})
        ax2_pie.set_xlim(0, 1); ax2_pie.set_ylim(0, 1)
        for at in autotexts:
            at.set_fontsize(7.5)
            at.set_color(WHITE)
            at.set_fontweight('bold')

        # 圖例：文字列表，最多 4 欄，畫在 axes 座標下方 25% 區域
        n_leg  = len(pie_labels)
        n_cols = min(4, n_leg)
        n_rows = (n_leg + n_cols - 1) // n_cols
        leg_fs = max(6.5, min(8.0, 60 / n_leg))
        leg_h  = 0.23 / max(n_rows, 1)   # 每列高度（axes 座標）
        col_w  = 1.0 / n_cols

        for idx, (lbl, col) in enumerate(zip(pie_labels, pie_colors)):
            row = idx // n_cols
            col_i = idx % n_cols
            xp = col_i * col_w + 0.01
            yp = 0.215 - row * leg_h
            # 色塊
            from matplotlib.patches import Rectangle as Rect
            ax2_pie.add_patch(Rect((xp, yp + leg_h*0.25), col_w*0.08, leg_h*0.5,
                facecolor=col, transform=ax2_pie.transAxes, clip_on=False))
            # 文字
            pct = pie_sizes[idx] / total_pie * 100
            ax2_pie.text(xp + col_w*0.11, yp + leg_h*0.50,
                f'{lbl} {pct:.1f}%',
                color=WHITE, fontsize=leg_fs, va='center', ha='left',
                transform=ax2_pie.transAxes, clip_on=False)
    else:
        ax2_pie.text(0.5, 0.5, '無持倉資料', color=GRAY, ha='center', va='center',
            transform=ax2_pie.transAxes)

    # ── P2 頁尾 ──
    fig2.text(0.50, 0.008, f'資料截至 {TODAY}  |  製表日期 {TODAY}  |  圖靈投資管理',
        color=GRAY, fontsize=7.5, ha='center', va='bottom')

    # ── 存檔（兩頁合併）──
    out_dir = os.path.expanduser('~/Desktop/Clients/Portfolio_Reports/')
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f'{name}_{TODAY_FILE}_資產現況.pdf')
    with PdfPages(path) as pdf:
        pdf.savefig(fig,  dpi=150, facecolor=BG)
        pdf.savefig(fig2, dpi=150, facecolor=BG)
    plt.close('all')

    te = time.time()
    print(f"[PDF] → {path}")
    print(f"[TIME] {te-t0:.1f}s")
    print(f"\n=== {name} 摘要 ===")
    print(f"期初: {i25:,.0f} | 餘額: {i27:,.0f} | 本期損益: {i29:+,.0f} ({period_ret:+.2f}%)")
    if cur:
        print(f"交易: {len(cur)}筆 | 勝率: {win_rate:.1f}% | 平均持有: {avg_days:.1f}天")
        if best and worst:
            print(f"最佳: {best['name']} {best['pnl']:+,.0f} | 最差: {worst['name']} {worst['pnl']:+,.0f}")
        if top3:
            print(f"Top3: {', '.join(f'{s}+{v:,.0f}' for s,v in top3)}")
    print(f"累計: {INIT:,.0f} → {i27:,.0f} ({cum_ret:+.2f}%)")

    return {
        'name': name, 'type': cfg['type'],
        'i25': i25, 'i27': i27, 'i29': i29,
        'init': INIT, 'period_ret': period_ret, 'cum_ret': cum_ret,
    }


# ── 主程式 ──
if __name__ == '__main__':
    if len(sys.argv) > 1 and sys.argv[1] != 'all':
        target = sys.argv[1]
        if target not in CLIENT_MAP:
            print(f"[ERROR] 未知客戶: {target}。可用: {list(CLIENT_MAP.keys())}")
            sys.exit(1)
        run_client(target)
    else:
        # 預設：全部客戶
        results = []
        for name in CLIENT_MAP:
            print(f"\n{'='*40}")
            print(f"處理：{name}")
            r = run_client(name)
            if r:
                results.append(r)

        if results:
            total_aum = sum(r['i27'] for r in results)
            print(f"\n\n{'='*50}")
            print(f"全客戶彙總（{TODAY}）")
            print(f"{'='*50}")
            print(f"{'客戶':<8} {'類型':<10} {'AUM':>12} {'佔比':>8} {'本期':>8} {'累計':>8}")
            print(f"{'-'*56}")
            for r in results:
                share = r['i27'] / total_aum * 100 if total_aum else 0
                print(f"{r['name']:<8} {r['type']:<10} {r['i27']:>12,.0f} "
                      f"{share:>7.1f}% {r['period_ret']:>+7.2f}% {r['cum_ret']:>+7.2f}%")
            print(f"{'-'*56}")
            w_period = sum(r['i29'] for r in results) / sum(r['i25'] for r in results) * 100
            print(f"{'合計':<8} {'':<10} {total_aum:>12,.0f} {'100.0%':>8} {w_period:>+7.2f}%")
