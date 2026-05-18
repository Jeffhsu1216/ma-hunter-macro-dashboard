"""
每日股價更新腳本 v4
執行時機：每日台灣收盤後（15:00+）
策略：yfinance 批次下載（一次 HTTP，所有股票 + TAIEX 同時取得）
  1. 掃描所有 Excel → 收集代號
  2. 全部加 .TW 後綴，一次 yf.download()
  3. 回傳 NaN 的代號 → 補一次 .TWO 批次
  4. 更新客戶 xlsx I 欄（新日期檔名）
  5. 更新 Jeff_Stock Analysis H 欄（覆寫原檔）
"""

import yfinance as yf
import openpyxl, os, glob, datetime

# ══════════════════════════════════════════
CLIENTS_DIR   = os.path.expanduser('~/Desktop/Clients')
EXCEL_DIR     = os.path.expanduser('~/Desktop/Clients/Excel')
JEFF_ANALYSIS = os.path.expanduser('~/Desktop/Stock Analysis/Jeff_Stock Analysis.xlsx')
TW = datetime.timezone(datetime.timedelta(hours=8))
# ══════════════════════════════════════════

def today_label():
    d = datetime.datetime.now(TW)
    return f"{d.month}/{d.day}收盤價", d.strftime('%Y%m%d')

# ── 掃描所有代號 ──────────────────────────
def collect_all_codes() -> set:
    codes = set()
    import re
    all_xlsx = [f for f in glob.glob(os.path.join(EXCEL_DIR, '*.xlsx'))
                if not os.path.basename(f).startswith('~$')
                and re.search(r'_\d{8}\.xlsx$', f)]
    for path in all_xlsx:
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            if 'Investment Portfolio' not in wb.sheetnames:
                continue
            ws = wb['Investment Portfolio']
            for r in range(4, 50):
                v = ws.cell(r, 4).value
                if v:
                    try:
                        ci = int(str(v).strip())
                        if 1000 <= ci <= 9999:
                            codes.add(str(ci))
                    except: pass
        except: pass
    try:
        wb = openpyxl.load_workbook(JEFF_ANALYSIS, data_only=True, read_only=True)
        monthly = sorted([s for s in wb.sheetnames if s.isdigit() and len(s) == 6])
        if monthly:
            ws = wb[monthly[-1]]
            for r in range(2, 50):
                v = ws.cell(r, 2).value
                if v:
                    try:
                        ci = int(str(v).strip())
                        if 1000 <= ci <= 9999:
                            codes.add(str(ci))
                    except: pass
    except: pass
    return codes

# ── yfinance 批次下載 ─────────────────────
def batch_download(tickers: list) -> dict:
    """一次下載，回傳 {ticker: close_price}，NaN 視為失敗"""
    if not tickers:
        return {}
    import pandas as pd
    df = yf.download(tickers, period='5d', progress=False,
                     auto_adjust=True, threads=True)
    if df.empty:
        return {}
    closes = df['Close'] if 'Close' in df.columns else df.xs('Close', axis=1, level=0)
    result = {}
    for t in (tickers if isinstance(tickers, list) else [tickers]):
        try:
            s = closes[t].dropna()
            if not s.empty:
                result[t] = round(float(s.iloc[-1]), 2)
        except: pass
    return result

def fetch_all_prices(codes: set) -> dict:
    """
    1. 全部試 .TW（上市）
    2. 失敗的補試 .TWO（上櫃）
    3. ^TWII = TAIEX 一起抓
    回傳 {'code': price, 'TAIEX': price}
    """
    tw_tickers  = [f"{c}.TW"  for c in sorted(codes)] + ['^TWII']
    print(f"📡 yfinance 批次下載 {len(tw_tickers)} 支（含TAIEX）...")
    tw_res = batch_download(tw_tickers)

    # 取 TAIEX
    taiex = tw_res.pop('^TWII', None)

    # 找失敗的（NaN）
    prices = {}
    failed = []
    for c in codes:
        t = f"{c}.TW"
        if t in tw_res:
            prices[c] = tw_res[t]
        else:
            failed.append(c)

    # 補試 .TWO
    if failed:
        print(f"  🔄 {len(failed)} 支試 .TWO：{failed}")
        two_tickers = [f"{c}.TWO" for c in failed]
        two_res = batch_download(two_tickers)
        still_fail = []
        for c in failed:
            t = f"{c}.TWO"
            if t in two_res:
                prices[c] = two_res[t]
                print(f"     {c}.TWO = {two_res[t]}")
            else:
                still_fail.append(c)
        if still_fail:
            print(f"  ⚠️  仍無法取得：{still_fail}")

    print(f"  ✅ 取得 {len(prices)} 支股票｜TAIEX = {taiex}")
    return prices, taiex

# ── 更新客戶 Excel ────────────────────────
def update_client_files(prices, label, date_str):
    import re
    all_files = [f for f in glob.glob(os.path.join(EXCEL_DIR, '*.xlsx'))
                 if not os.path.basename(f).startswith('~$')
                 and re.search(r'_\d{8}\.xlsx$', f)]
    # 每位客戶只取最新檔
    latest = {}
    for f in all_files:
        base = os.path.basename(f)
        parts = base.rsplit('_', 1)
        client = parts[0] if (len(parts)==2 and parts[1].replace('.xlsx','').isdigit()) else base.replace('.xlsx','')
        if client not in latest or f > latest[client]:
            latest[client] = f

    for client_name, path in sorted(latest.items()):
        try:
            wb = openpyxl.load_workbook(path)
            if 'Investment Portfolio' not in wb.sheetnames:
                continue
            ws = wb['Investment Portfolio']
            ws.cell(3, 9).value = label          # I3 標題
            updated = 0
            for r in range(4, ws.max_row + 1):
                v = ws.cell(r, 4).value
                if not v: continue
                try:
                    ci = int(str(v).strip())
                    if not (1000 <= ci <= 9999): continue
                except: continue
                if str(ci) in prices:
                    ws.cell(r, 9).value = prices[str(ci)]
                    updated += 1
            new_path = os.path.join(EXCEL_DIR, f"{client_name}_{date_str}.xlsx")
            wb.save(new_path)
            # 舊檔改名（刪除），只保留今日日期版本
            if os.path.abspath(path) != os.path.abspath(new_path):
                os.remove(path)
            print(f"  ✅ {client_name}：{updated} 支持股 → {client_name}_{date_str}.xlsx（舊檔已移除）")
        except Exception as e:
            print(f"  ❌ {client_name}：{e}")

# ── 更新 Jeff_Stock Analysis ──────────────
def update_jeff_analysis(prices, taiex, label):
    try:
        wb = openpyxl.load_workbook(JEFF_ANALYSIS)
        monthly = sorted([s for s in wb.sheetnames if s.isdigit() and len(s) == 6])
        ws = wb[monthly[-1]]
        ws.cell(1, 8).value = label              # H1 標題
        updated = 0
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 2).value
            if not v: continue
            try:
                ci = int(str(v).strip())
                if not (1000 <= ci <= 9999): continue
            except: continue
            if str(ci) in prices:
                ws.cell(r, 8).value = prices[str(ci)]
                updated += 1
        if taiex:
            ws.cell(44, 8).value = taiex
        wb.save(JEFF_ANALYSIS)
        print(f"  ✅ Jeff_Stock Analysis [{monthly[-1]}]：{updated} 支｜TAIEX={taiex}")
    except Exception as e:
        print(f"  ❌ Jeff_Stock Analysis：{e}")

# ── 主流程 ────────────────────────────────
def run():
    label, date_str = today_label()
    print(f"\n📅 每日股價更新 v3 | {label}")
    print("=" * 45)

    codes = collect_all_codes()
    print(f"📋 掃描到 {len(codes)} 個代號")

    prices, taiex = fetch_all_prices(codes)

    print("\n【客戶 Excel】")
    update_client_files(prices, label, date_str)

    print("\n【Jeff_Stock Analysis】")
    update_jeff_analysis(prices, taiex, label)

    print("\n✅ 完成")

if __name__ == '__main__':
    run()
